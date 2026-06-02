#!/usr/bin/env python3
"""
GAMALI PATBİS — Python Bridge v2.2.0
=====================================
Tauri ↔ Python iletişim köprüsü (JSON-RPC over stdin/stdout).
Komutlar:
  saglik           → durum kontrolü
  db.depolar       → depo listesi
  txt.bul          → XML'e karşılık TXT dosyasını bul
  xml.parse        → XML SummaryItem özeti + gonderi bilgi
  txt.parse        → TXT ürünleri parse + XML eşleşme
  stok.kaydet      → veritabanına kayıt
"""

import sys
import json
import os
import re
import shutil
import sqlite3
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

# Add script directory to sys.path to support imports in embedded environments
script_dir = Path(__file__).parent.resolve()
if str(script_dir) not in sys.path:
    sys.path.insert(0, str(script_dir))

from core import drive_manager

VERSIYON = "2.3.0"


# ── Startup mesajı ──
def startup():
    """Bridge hazır sinyali gönder."""
    print(json.dumps({"ready": True, "versiyon": VERSIYON}), flush=True)


# ── Komut işleyici ──
def handle(req):
    """Gelen JSON-RPC isteğini işle, sonuç döndür."""
    req_id = req.get("id", 0)
    method = req.get("method", "")
    params = req.get("params", {})

    try:
        result = dispatch(method, params)
        return {"id": req_id, "result": result, "error": None}
    except Exception as e:
        return {"id": req_id, "result": None, "error": str(e)}


def dispatch(method, params):
    """Method router."""

    # ── Sağlık ──
    if method == "saglik":
        return {
            "durum": "OK",
            "uygulama": "GAMALI PATBİS Python Core",
            "versiyon": VERSIYON,
        }

    # ── Depo Listesi ──
    elif method == "db.depolar":
        return _depolar_getir()

    # ── Depo Özeti (Dashboard) ──
    elif method == "db.depo_ozeti":
        return _depo_ozeti_getir()

    # ── Son Yüklemeler (Dashboard) ──
    elif method == "db.son_yuklemeler":
        return _son_yuklemeler_getir(params.get("limit", 20))

    # ── Barkod Sorgula ──
    elif method == "barkod.sorgula":
        return _barkod_sorgula(params.get("uid", ""))

    # ── Personel Yönetimi ──
    elif method == "db.personeller":
        return _personeller_getir()
    elif method == "db.personel_ekle":
        return _personel_ekle(params)
    elif method == "db.personel_guncelle":
        return _personel_guncelle(params)
    elif method == "db.personel_sil":
        return _personel_sil(params)

    # ── TXT Otomatik Bul ──
    elif method == "txt.bul":
        return _txt_bul(params.get("xml_yolu", ""))

    # ── XML Parse — SummaryItem Özeti ──
    elif method == "xml.parse":
        return _xml_parse(params.get("dosya_yolu", ""))

    # ── TXT Parse — Ürünler + XML Eşleşme ──
    elif method == "txt.parse":
        return _txt_parse(
            params.get("txt_yolu", ""),
            params.get("xml_yolu", ""),
        )

    # ── Stok Kaydet ──
    elif method == "stok.kaydet":
        return _stok_kaydet(params)

    # ── Stok Özet ──
    elif method == "stok.ozet":
        return _stok_ozet(params)

    # ── Stok Yönetimi ──
    elif method == "stok.sifirla":
        return _stok_sifirla(params)

    # ── Google Drive ──
    elif method == "drive.durum":
        return _drive_durum()
    elif method == "drive.baglan":
        return _drive_baglan()
    elif method == "drive.yedekle":
        return _drive_yedekle()
    elif method == "drive.dosyalar":
        return _drive_dosyalar()
    elif method == "drive.login":
        return _drive_login(params)

    elif method == "stok.veritabani_sil":
        return _stok_veritabani_sil()

    elif method == "stok.cikis_onizle":
        return _stok_cikis_onizle(params)

    elif method == "stok.cikis":
        return _stok_cikis(params)

    elif method == "stok.envanter":
        return _stok_envanter(params)

    elif method == "stok.sayim_karsilastir":
        return _stok_sayim_karsilastir(params)

    # ── Raporlar ──
    elif method == "rapor.veri":
        return _rapor_veri(params)

    elif method == "rapor.olustur":
        return _rapor_olustur(params)

    # ── Dosya İşlemleri ──
    elif method == "dosya.kaydet":
        return _dosya_kaydet(params)

    # ── Online Sync & Güvenlik Komutları ──
    elif method == "db.islem_gunlugu":
        return _islem_gunlugu_getir(params)
    elif method == "db.urun_sorunlari":
        return _urun_sorunlari_getir(params)
    elif method == "db.sunucu_durumu":
        return _sunucu_durumu_getir()

    else:
        raise ValueError(f"Bilinmeyen komut: {method}")


# ══════════════════════════════════════════════════════════════════════════════
# DEPO
# ══════════════════════════════════════════════════════════════════════════════

def _depolar_getir():
    """SQLite'tan depo listesi çek. Eksik sabit depoları DB'ye ekle."""
    try:
        conn = _db_baglanti()
        
        # Sabit depoları kontrol et ve eksikse ekle
        sabit = _sabit_depolar()
        for d in sabit:
            conn.execute(
                "INSERT OR IGNORE INTO depolar (kod, ad) VALUES (?, ?)",
                (d["kod"], d["ad"]),
            )
        conn.commit()
        
        rows = conn.execute("SELECT id, kod, ad FROM depolar WHERE deleted_at IS NULL ORDER BY ad").fetchall()
        result = [dict(r) for r in rows]
        conn.close()
        return result if result else sabit
    except Exception:
        return _sabit_depolar()


def _sabit_depolar():
    return [
        {"kod": "33-2019-00419", "ad": "Özaltın 33-2019-00419"},
        {"kod": "33-2019-00422", "ad": "Özaltın 33-2019-00422"},
        {"kod": "33-2019-00425", "ad": "Özaltın 33-2019-00425"},
        {"kod": "33-2019-00428", "ad": "Özaltın 33-2019-00428"},
        {"kod": "33-2019-00454", "ad": "Özaltın 33-2019-00454"},
        {"kod": "33-2019-00457", "ad": "Özaltın 33-2019-00457"},
    ]


def _depo_ozeti_getir():
    """
    Her depo için toplam stok sayısını döndür.
    Döndürülen format: [{depo_kod, depo_adi, stokta, palet_sayisi, kutu_sayisi}]
    """
    try:
        conn = _db_baglanti()
        # Depo bazlı toplam ürün, ağırlık, yükleme sayısı
        rows = conn.execute("""
            SELECT
                d.kod          AS depo_kod,
                d.ad           AS depo_adi,
                COALESCE(SUM(y.urun_sayisi), 0)    AS stokta,
                COALESCE(SUM(y.cesit_sayisi), 0)   AS palet_sayisi,
                COUNT(y.id)                         AS kutu_sayisi
            FROM depolar d
            LEFT JOIN stok_yuklemeler y
                ON y.depo_id = d.id AND y.deleted_at IS NULL
            WHERE d.deleted_at IS NULL
            GROUP BY d.id, d.kod, d.ad
            ORDER BY d.ad
        """).fetchall()
        conn.close()
        return [dict(r) for r in rows]
    except Exception:
        # DB yoksa sıfır değerli depo listesi döndür
        return [
            {"depo_kod": d["kod"], "depo_adi": d["ad"],
             "stokta": 0, "palet_sayisi": 0, "kutu_sayisi": 0}
            for d in _sabit_depolar()
        ]


def _son_yuklemeler_getir(limit=20):
    """
    Son yapılan yüklemeleri döndür — Dashboard tablosu için.
    Döndürülen format: [{depo_kod, depo_ad, dosya_adi, tarih, urun_sayisi}]
    """
    try:
        conn = _db_baglanti()
        rows = conn.execute("""
            SELECT
                d.kod                       AS depo_kod,
                y.depo_adi                  AS depo_ad,
                COALESCE(
                    CASE
                        WHEN y.xml_dosya != '' THEN
                            REPLACE(REPLACE(y.xml_dosya, '\\', '/'), '\\\\', '/')
                        ELSE ''
                    END,
                    ''
                )                           AS dosya_tam,
                y.gonderi_no                AS gonderi_no,
                SUBSTR(y.yukleme_tarihi, 1, 16) AS tarih,
                y.urun_sayisi,
                y.cesit_sayisi,
                y.toplam_agirlik
            FROM stok_yuklemeler y
            LEFT JOIN depolar d ON d.id = y.depo_id
            WHERE y.deleted_at IS NULL
            ORDER BY y.yukleme_tarihi DESC
            LIMIT ?
        """, (limit,)).fetchall()
        conn.close()

        sonuc = []
        for r in rows:
            r = dict(r)
            # Dosya adını yoldan çıkar
            tam = r.get("dosya_tam", "")
            dosya_adi = tam.split("/")[-1] if tam else (r.get("gonderi_no") or "-")
            sonuc.append({
                "depo_kod": r["depo_kod"] or "-",
                "depo_ad": r["depo_ad"] or "-",
                "dosya_adi": dosya_adi,
                "tarih": (r["tarih"] or "-").replace("T", " "),
                "urun_sayisi": r["urun_sayisi"] or 0,
                "cesit_sayisi": r["cesit_sayisi"] or 0,
                "toplam_agirlik": r["toplam_agirlik"] or 0,
            })
        return sonuc
    except Exception:
        return []


def _barkod_sorgula(uid):
    """
    UID'ye göre hiyerarşik ağaç yapısı döndürür (Palet -> Kutu -> Ürün).
    """
    if not uid:
        return {"basarili": False, "mesaj": "Barkod girmediniz."}

    try:
        conn = _db_baglanti()
        
        # 1. Kökü belirle (Kullanıcı İsteği: Ürün aranırsa sadece ürün, koliyse sadece koli ağacı)
        root_uid = uid

        # 2. Ağacı oluştur
        def build_tree(current_uid):
            # Birim mi?
            unit = conn.execute("""
                SELECT b.*, y.depo_adi, d.kod as depo_kod, y.gonderi_no
                FROM birimler b
                JOIN stok_yuklemeler y ON y.id = b.yukleme_id
                LEFT JOIN depolar d ON d.id = y.depo_id
                WHERE b.uid = ?
            """, (current_uid,)).fetchone()

            if unit:
                tip = "palet" if unit["seviye"] == "04" else "kutu"
                
                # Çocuk birimler (Sub-units)
                sub_units = conn.execute("SELECT uid FROM birimler WHERE parent_uid = ?", (current_uid,)).fetchall()
                # Çocuk ürünler (Items)
                items = conn.execute("""
                    SELECT u.*, y.depo_adi, d.kod as depo_kod, y.gonderi_no
                    FROM urunler u
                    JOIN stok_yuklemeler y ON y.id = u.yukleme_id
                    LEFT JOIN depolar d ON d.id = y.depo_id
                    WHERE u.parent_uid = ? AND u.deleted_at IS NULL
                """, (current_uid,)).fetchall()

                cocuklar = []
                for su in sub_units:
                    cocuklar.append(build_tree(su["uid"]))
                
                for it in items:
                    # Ürünler yapraktır, ama tip olarak 'urun' dönerler
                    cocuklar.append({
                        "tip": "urun",
                        "uid": it["uid"],
                        "psn": it["psn"],
                        "sid": it["sid"],
                        "durum": "aktif",
                        "depo": it["depo_adi"],
                        "urun_adi": it["urun_adi"],
                        "urun_kodu": it["kod"],
                        "parti_no": it["lot"],
                        "agirlik": f"{it['agirlik']:.3f} kg",
                        "uretim_tarihi": it["skt"], # Örnekleme için SKT/Lot kullanılabilir
                        "son_kullanma": it["skt"],
                        "ust_kutu": it["parent_uid"],
                        "gonderi_no": it["gonderi_no"]
                    })

                return {
                    "tip": tip,
                    "uid": unit["uid"],
                    "psn": "-", # Birimlerde PSN genellikle sid ile aynı veya opsiyonel
                    "sid": "-",
                    "durum": "aktif",
                    "depo": unit["depo_adi"],
                    "seviye": unit["seviye"],
                    "parent_uid": unit["parent_uid"],
                    "gonderi_no": unit["gonderi_no"],
                    "cocuklar": cocuklar
                }
            
            # Ürün mü? (Eğer sorgulanan şey direkt bir ürünse ve yukarısı yoksa)
            item = conn.execute("""
                SELECT u.*, y.depo_adi, d.kod as depo_kod, y.gonderi_no
                FROM urunler u
                JOIN stok_yuklemeler y ON y.id = u.yukleme_id
                LEFT JOIN depolar d ON d.id = y.depo_id
                WHERE u.uid = ? AND u.deleted_at IS NULL
            """, (current_uid,)).fetchone()
            
            if item:
                parent_pallet_uid = None
                if item["parent_uid"]:
                    p_box = conn.execute("SELECT parent_uid FROM birimler WHERE uid = ?", (item["parent_uid"],)).fetchone()
                    if p_box:
                        parent_pallet_uid = p_box["parent_uid"]

                return {
                    "tip": "urun",
                    "uid": item["uid"],
                    "psn": item["psn"],
                    "sid": item["sid"],
                    "durum": "aktif",
                    "depo": item["depo_adi"],
                    "urun_adi": item["urun_adi"],
                    "urun_kodu": item["kod"],
                    "parti_no": item["lot"],
                    "agirlik": f"{item['agirlik']:.3f} kg",
                    "son_kullanma": item["skt"],
                    "ust_kutu": item["parent_uid"],
                    "ust_palet": parent_pallet_uid,
                    "gonderi_no": item["gonderi_no"]
                }
            
            return None

        tree_data = build_tree(root_uid)
        
        # 3. Lot / Parti No ile ara (Eğer ağaç bulunamadıysa lot olarak ara)
        if not tree_data:
            lot_urunleri = conn.execute("""
                SELECT u.*, y.depo_adi, y.gonderi_no, d.kod as depo_kod
                FROM urunler u
                JOIN stok_yuklemeler y ON y.id = u.yukleme_id
                LEFT JOIN depolar d ON d.id = y.depo_id
                WHERE u.lot = ? AND u.deleted_at IS NULL
            """, (uid,)).fetchall()

            if lot_urunleri:
                urunler = [dict(r) for r in lot_urunleri]
                ilk = urunler[0]
                tree_data = {
                    "tip": "kutu",
                    "uid": uid,
                    "psn": "-",
                    "sid": uid,
                    "durum": "aktif",
                    "depo": ilk["depo_adi"],
                    "urun_adi": ilk["urun_adi"],
                    "parti_no": ilk["lot"],
                    "son_kullanma": ilk["skt"],
                    "seviye": "Lot Grubu",
                    "cocuklar": [
                        {
                            "tip": "urun",
                            "uid": it["uid"],
                            "psn": it["psn"],
                            "sid": it["sid"],
                            "durum": "aktif",
                            "depo": it["depo_adi"],
                            "urun_adi": it["urun_adi"],
                            "urun_kodu": it["kod"],
                            "parti_no": it["lot"],
                            "agirlik": f"{it['agirlik']:.3f} kg",
                            "uretim_tarihi": it["skt"],
                            "son_kullanma": it["skt"],
                            "ust_kutu": it["parent_uid"] or uid
                        } for it in urunler
                    ]
                }

        conn.close()

        if not tree_data:
            return {"basarili": False, "mesaj": "Barkod bulunamadı."}
            
        return {"basarili": True, "data": tree_data}

    except Exception as e:
        return {"basarili": False, "mesaj": f"Sorgulama hatası: {e}"}


def _db_yolu():
    """
    Veritabanı dosya yolunu belirle.
    Öncelik: PATBIS_DB env → %APPDATA%/gamali-patbis/DB/gamalipatbis.db (Gelişmiş veri güvenliği ve yazma izni için)
    Eğer APPDATA içinde yoksa ve uygulama dizininde (resources) hazır db varsa ilk çalıştırmada oraya kopyalar.
    """
    env_yol = os.environ.get("PATBIS_DB")
    if env_yol:
        return env_yol

    # 1. Gömülü/Kaynak veritabanı (App directory - resources)
    base = os.path.dirname(os.path.abspath(__file__))
    bundled_db = os.path.join(base, "DB", "gamalipatbis.db")

    # 2. Güvenli Yazılabilir Veritabanı Yolu (%APPDATA% veya Home)
    appdata = os.environ.get("APPDATA")
    if appdata:
        db_dir = os.path.join(appdata, "gamali-patbis", "DB")
    else:
        home = os.path.expanduser("~")
        db_dir = os.path.join(home, ".gamali-patbis", "DB")

    target_db = os.path.join(db_dir, "gamalipatbis.db")

    # İlk kez çalışıyorsa ve henüz yazılabilir konumda DB yoksa, varsa gömülüyü kopyalayalım
    if not os.path.exists(target_db):
        try:
            os.makedirs(db_dir, exist_ok=True)
            if os.path.exists(bundled_db):
                shutil.copy2(bundled_db, target_db)
        except Exception:
            # Kopyalama veya klasör oluşturma başarısız olursa
            pass

    try:
        os.makedirs(db_dir, exist_ok=True)
        return target_db
    except Exception:
        # Son çare fallback (geçici veya dev ortamı)
        try:
            os.makedirs(os.path.join(base, "DB"), exist_ok=True)
        except Exception:
            pass
        return bundled_db


def _yedek_dizini():
    """Dosya yedekleri için dizin: DB/yedekler/"""
    db_dir = os.path.dirname(_db_yolu())
    yedek_dir = os.path.join(db_dir, "yedekler")
    os.makedirs(yedek_dir, exist_ok=True)
    return yedek_dir


# ══════════════════════════════════════════════════════════════════════════════
# TXT OTOMATIK BUL
# ══════════════════════════════════════════════════════════════════════════════

def _txt_bul(xml_yolu):
    """XML ile aynı klasörde TXT dosyasını bul."""
    if not xml_yolu:
        return {"bulundu": False, "mesaj": "XML yolu boş"}

    xml_path = Path(xml_yolu)
    if not xml_path.exists():
        return {"bulundu": False, "mesaj": "XML dosyası bulunamadı"}

    # 1. Aynı isim .txt
    eslesme = xml_path.with_suffix(".txt")
    if eslesme.exists():
        return {"bulundu": True, "yol": str(eslesme), "tip": "otomatik"}

    # 2. Klasördeki tek TXT
    txt_dosyalar = list(xml_path.parent.glob("*.txt"))
    if len(txt_dosyalar) == 1:
        return {"bulundu": True, "yol": str(txt_dosyalar[0]), "tip": "otomatik"}

    # 3. İsim benzerliği
    if txt_dosyalar:
        stem = xml_path.stem.lower()
        eslesmeler = sorted(
            txt_dosyalar,
            key=lambda p: sum(c in p.stem.lower() for c in stem),
            reverse=True,
        )
        return {"bulundu": True, "yol": str(eslesmeler[0]), "tip": "benzer"}

    return {"bulundu": False, "mesaj": "TXT bulunamadı"}


# ══════════════════════════════════════════════════════════════════════════════
# XML PARSE — SummaryItem Özeti
# ══════════════════════════════════════════════════════════════════════════════

def _xml_parse(dosya_yolu):
    """
    FEEM XML dosyasını parse et.
    Döndürdüğü veri:
      - gonderi_bilgi: mesaj_id, gonderi_no, teslim_tarihi, gonderen, alici
      - summary_items: [{sid, psn, urun_kodu, urun_adi, lot, skt, uretim_tarihi, adet}]
      - toplam_item: XML'deki toplam <Item> sayısı
      - toplam_unit: XML'deki toplam <Unit> sayısı (koliler)
    """
    if not dosya_yolu or not os.path.exists(dosya_yolu):
        return {"basarili": False, "mesaj": f"Dosya bulunamadı: {dosya_yolu}"}

    try:
        tree = ET.parse(dosya_yolu)
        root = tree.getroot()
    except ET.ParseError as e:
        return {"basarili": False, "mesaj": f"XML parse hatası: {e}"}

    # ── Gönderi Bilgisi ──
    gonderi_bilgi = {
        "mesaj_id": _text(root, "MessageID"),
        "gonderi_no": _text(root, "ShipmentNumber"),
        "teslim_tarihi": _tarih_temizle(_text(root, "ExpectedDeliveryDate")),
        "gonderen": _text(root, ".//Sender/Name"),
        "gonderen_kodu": _text(root, ".//Sender/Code"),
        "alici": _text(root, ".//Receiver/Name"),
        "alici_kodu": _text(root, ".//Receiver/Code"),
    }

    # ── SummaryItem'lar — Ürün Çeşitleri ──
    summary_items = []
    sid_adet = {}  # SID → adet (Item sayısından hesaplanacak)

    # Önce tüm Item'ları say — SID bazlı
    tum_itemler = root.findall(".//Item")
    for item in tum_itemler:
        sid = item.get("SID", "")
        sid_adet[sid] = sid_adet.get(sid, 0) + 1

    # SummaryItem'ları oku
    for si in root.findall(".//SummaryItem"):
        sid = si.get("SID", "")
        psn = si.get("PSN", "")
        summary_items.append({
            "sid": sid,
            "psn": psn,
            "urun_kodu": _text(si, "ProducerProductCode"),
            "urun_adi": _text(si, "ProducerProductName"),
            "lot": _text(si, "BatchNumber"),
            "skt": _tarih_temizle(_text(si, "ExpiryDate")),
            "uretim_tarihi": _tarih_temizle(_text(si, "ProductionDate")),
            "adet": sid_adet.get(sid, 0),
        })

    # ── Toplam Unit sayısı (PL=02 seviyesi — koliler) ──
    tum_unitler = root.findall(".//Unit")
    toplam_unit = len(tum_unitler)

    # ── Hiyerarşi ve Birim Detayları ──
    # Parent haritası: Child UID -> Parent UID
    parent_map = {}
    unit_details = []  # [{uid, seviye, adet, parent_uid}]

    # Element -> Parent Element haritası (tüm ağaç için)
    elt_parent = {c: p for p in root.iter() for c in p}

    for unit in tum_unitler:
        unit_uid = unit.get("UID", "")
        if not unit_uid: continue
        
        seviye = _text(unit, "PackagingLevel")
        items = unit.findall(".//Item")
        
        # Bu ünitenin bir üst ünitesi var mı?
        parent_unit_uid = None
        curr = elt_parent.get(unit)
        while curr is not None:
            if curr.tag == "Unit":
                parent_unit_uid = curr.get("UID", "")
                break
            curr = elt_parent.get(curr)

        unit_details.append({
            "uid": unit_uid,
            "seviye": seviye,
            "adet": len(items),
            "parent_uid": parent_unit_uid
        })

        # Item'ların parent'ı bu ünitedir
        for item in items:
            item_uid = item.get("UID", "")
            if item_uid:
                parent_map[item_uid] = unit_uid

    return {
        "basarili": True,
        "gonderi_bilgi": gonderi_bilgi,
        "summary_items": summary_items,
        "toplam_item": len(tum_itemler),
        "toplam_unit": toplam_unit,
        "parent_map": parent_map,
        "unit_details": unit_details
    }


# ══════════════════════════════════════════════════════════════════════════════
# TXT PARSE — Ürünler + XML Eşleşme
# ══════════════════════════════════════════════════════════════════════════════

def _txt_parse(txt_yolu, xml_yolu):
    """
    CK65 barkod TXT dosyasını parse et.
    Desteklenen formatlar:
      1. GS1-128 parantezli: (90)PSN(250)UID(240)kod(20)xx(3103)ağırlık
      2. Elle girilen parantezli: (250)UID
      3. Yalın alfanümerik UID: 21031318347

    Android APK başlık satırlarını (Depo:, Kullanici:, vb.) otomatik filtreler.
    XML'deki Item UID'leri ile karşılaştırarak eşleşme + meta veri hydration yapar.
    """
    if not txt_yolu or not os.path.exists(txt_yolu):
        return {"basarili": False, "mesaj": f"TXT dosyası bulunamadı: {txt_yolu}"}

    # ── XML'den haritalar oluştur ──
    xml_uid_sid = {}   # UID → SID
    xml_uid_psn = {}   # UID → PSN
    xml_sid_kod = {}   # SID → ProducerProductCode (ilk 3 hane)
    xml_unit_items = {} # unit_uid → list of Item elements (to support unpacking boxes)

    if xml_yolu and os.path.exists(xml_yolu):
        try:
            tree = ET.parse(xml_yolu)
            root = tree.getroot()

            # SummaryItem'lardan SID → ürün kodu haritası
            for si in root.findall(".//SummaryItem"):
                sid = si.get("SID", "")
                prod_code_full = _text(si, "ProducerProductCode")
                prod_code = prod_code_full[:3] if prod_code_full else ""
                if sid:
                    xml_sid_kod[sid] = prod_code

            # Item'lardan UID → SID, PSN haritası
            for item in root.findall(".//Item"):
                uid = item.get("UID", "")
                sid = item.get("SID", "")
                psn = item.get("PSN", "")
                if uid:
                    xml_uid_sid[uid] = sid
                    if psn:
                        xml_uid_psn[uid] = psn

            # Unit'lerden Unit_UID -> [Items] haritası (koli açma)
            for unit in root.findall(".//Unit"):
                unit_uid = unit.get("UID", "")
                if unit_uid:
                    xml_unit_items[unit_uid] = unit.findall(".//Item")
        except Exception:
            pass  # XML okunamazsa eşleşme yapılmaz

    # ── Android APK başlık satırlarını atlama deseni ──
    header_pattern = re.compile(
        r"^(depo|kullanici|kullanim[_ ]?yeri|kullanim|tarih|adet|gs1|manuel|barcode|barkod)\s*:",
        re.IGNORECASE
    )

    # ── TXT satırlarını parse et ──
    gs1_pattern = re.compile(
        r"\(90\)([^(]+)"       # PSN
        r"\(250\)([^(]+)"      # UID
        r"\(240\)([^(]+)"      # Kod (SID referansı)
        r"\(20\)([^(]+)"       # Paketleme seviyesi
        r"(?:\(30\)\d+)?"      # Opsiyonel adet
        r"\(3[01]\d{2}\)(\d+)" # Ağırlık (3103 veya 3101 vb.)
    )

    # Elle girilen (250)UID deseni
    manual_uid_pattern = re.compile(r"^\(250\)([A-Za-z0-9_-]+)$")

    # Yalın alfanümerik UID deseni (5-30 karakter)
    raw_uid_pattern = re.compile(r"^[A-Za-z0-9_-]{5,30}$")

    urunler = []
    eklenen_uidler = set()
    eslesen = 0
    eslesmeyen = 0
    hatalar = []

    with open(txt_yolu, "r", encoding="utf-8-sig", errors="replace") as f:
        for satir_no, satir in enumerate(f, 1):
            satir = satir.strip()
            if not satir:
                continue

            # ── 0. Android APK başlık satırlarını atla ──
            satir_lower = satir.lower()
            if (header_pattern.match(satir_lower) or
                satir_lower in ("barcode", "barkod")):
                continue

            # Ayırıcı çizgileri atla (--- veya ===)
            if re.match(r"^[-=]{3,}$", satir):
                continue

            # ── 1. Standart GS1-128 formatı ──
            m = gs1_pattern.search(satir)
            if m:
                psn = m.group(1).strip()
                uid = m.group(2).strip()
                kod = m.group(3).strip()
                paketleme = m.group(4).strip()
                agirlik_ham = m.group(5).strip()

                try:
                    agirlik = int(agirlik_ham) / 1000.0
                    agirlik_str = f"{agirlik:.3f}"
                except ValueError:
                    agirlik_str = agirlik_ham

                # Eğer bu UID bir Unit (Koli/Kutu) ise altındaki ürünleri açımla
                if uid in xml_unit_items:
                    for item in xml_unit_items[uid]:
                        item_uid = item.get("UID", "")
                        if item_uid in eklenen_uidler:
                            continue
                        eklenen_uidler.add(item_uid)

                        item_sid = item.get("SID", "")
                        item_psn = item.get("PSN", "")
                        item_kod = xml_sid_kod.get(item_sid, "000")

                        eslesen += 1
                        urunler.append({
                            "uid": item_uid,
                            "psn": item_psn,
                            "sid": item_sid,
                            "kod": item_kod,
                            "paketleme": "01",  # Ürün seviyesi
                            "agirlik": "0.000",
                            "xml_eslesme": True,
                            "koli_uid": uid
                        })
                    continue

                if uid in eklenen_uidler:
                    continue
                eklenen_uidler.add(uid)

                xml_eslesti = uid in xml_uid_sid
                sid = xml_uid_sid.get(uid, kod)

                if xml_eslesti:
                    eslesen += 1
                else:
                    eslesmeyen += 1

                urunler.append({
                    "uid": uid,
                    "psn": psn,
                    "sid": sid,
                    "kod": kod,
                    "paketleme": paketleme,
                    "agirlik": agirlik_str,
                    "xml_eslesme": xml_eslesti,
                })
                continue

            # ── 2. Elle girilen parantezli UID: (250)ABC12345 ──
            m_manual = manual_uid_pattern.match(satir)
            if m_manual:
                uid = m_manual.group(1).strip()
            # ── 3. Yalın alfanümerik UID: 21031318347 ──
            elif raw_uid_pattern.match(satir):
                uid = satir
            else:
                hatalar.append(f"Satır {satir_no}: Parse edilemedi → {satir[:40]}")
                continue

            # Eğer bu UID bir Unit (Koli/Kutu) ise altındaki ürünleri açımla
            if uid in xml_unit_items:
                for item in xml_unit_items[uid]:
                    item_uid = item.get("UID", "")
                    if item_uid in eklenen_uidler:
                        continue
                    eklenen_uidler.add(item_uid)

                    item_sid = item.get("SID", "")
                    item_psn = item.get("PSN", "")
                    item_kod = xml_sid_kod.get(item_sid, "000")

                    eslesen += 1
                    urunler.append({
                        "uid": item_uid,
                        "psn": item_psn,
                        "sid": item_sid,
                        "kod": item_kod,
                        "paketleme": "01",
                        "agirlik": "0.000",
                        "xml_eslesme": True,
                        "koli_uid": uid
                    })
                continue

            if uid in eklenen_uidler:
                continue
            eklenen_uidler.add(uid)

            # ── XML Hydration: Elle girilen UID'yi XML'de ara ──
            xml_eslesti = uid in xml_uid_sid
            if xml_eslesti:
                sid = xml_uid_sid[uid]
                psn = xml_uid_psn.get(uid, "")
                kod = xml_sid_kod.get(sid, "000")
                eslesen += 1
            else:
                sid = ""
                psn = ""
                kod = "000"
                eslesmeyen += 1

            urunler.append({
                "uid": uid,
                "psn": psn,
                "sid": sid,
                "kod": kod,
                "paketleme": "00",
                "agirlik": "0.000",
                "xml_eslesme": xml_eslesti,
            })

    return {
        "basarili": True,
        "urunler": urunler,
        "toplam": len(urunler),
        "eslesen": eslesen,
        "eslesmeyen": eslesmeyen,
        "parse_hatalari": len(hatalar),
        "hatalar_listesi": hatalar,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STOK KAYDET — Tam kapsamlı (DB + tekil ürünler + dosya yedekleme)
# ══════════════════════════════════════════════════════════════════════════════

def _db_baglanti():
    """SQLite bağlantısı aç, tabloları oluştur."""
    db_yolu = _db_yolu()
    conn = sqlite3.connect(db_yolu)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    # ── Tablo Şeması ──

    # 1. Depolar
    conn.execute("""
        CREATE TABLE IF NOT EXISTS depolar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kod TEXT NOT NULL UNIQUE,
            ad TEXT NOT NULL,
            deleted_at TEXT
        )
    """)

    # 2. Stok Yüklemeleri (ana kayıt — her XML+TXT yükleme bir satır)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS stok_yuklemeler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            depo_id INTEGER,
            depo_adi TEXT,
            gonderi_no TEXT,
            mesaj_id TEXT,
            gonderen TEXT,
            alici TEXT,
            teslim_tarihi TEXT,
            xml_dosya TEXT,
            txt_dosya TEXT,
            xml_yedek TEXT,
            txt_yedek TEXT,
            urun_sayisi INTEGER DEFAULT 0,
            cesit_sayisi INTEGER DEFAULT 0,
            toplam_agirlik REAL DEFAULT 0,
            yukleme_tarihi TEXT NOT NULL,
            deleted_at TEXT
        )
    """)

    # 3. Ürün Çeşitleri (SummaryItem — her yükleme için XML'den gelen çeşitler)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS urun_cesitleri (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            yukleme_id INTEGER NOT NULL,
            sid TEXT,
            psn TEXT,
            urun_kodu TEXT,
            urun_adi TEXT,
            lot TEXT,
            skt TEXT,
            uretim_tarihi TEXT,
            adet INTEGER DEFAULT 0,
            FOREIGN KEY (yukleme_id) REFERENCES stok_yuklemeler(id)
        )
    """)

    # 4. Ürünler (tekil — her TXT satırı bir ürün)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS urunler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            yukleme_id INTEGER NOT NULL,
            uid TEXT NOT NULL,
            psn TEXT,
            sid TEXT,
            kod TEXT,
            paketleme TEXT,
            agirlik REAL DEFAULT 0,
            xml_eslesme INTEGER DEFAULT 0,
            urun_adi TEXT,
            lot TEXT,
            skt TEXT,
            deleted_at TEXT,
            parent_uid TEXT,
            FOREIGN KEY (yukleme_id) REFERENCES stok_yuklemeler(id)
        )
    """)

    # 5. Birimler (Koli/Kutu detayları)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS birimler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            yukleme_id INTEGER NOT NULL,
            uid TEXT NOT NULL UNIQUE,
            seviye TEXT,
            adet INTEGER DEFAULT 0,
            parent_uid TEXT,
            FOREIGN KEY (yukleme_id) REFERENCES stok_yuklemeler(id)
        )
    """)

    # 6. Personeller
    conn.execute("""
        CREATE TABLE IF NOT EXISTS personeller (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ad_soyad TEXT NOT NULL,
            unvan TEXT,
            aktif INTEGER DEFAULT 1,
            olusturma TEXT DEFAULT (datetime('now'))
        )
    """)

    # 1. İşlem Günlüğü (Audit Log) - Tamper-Proof
    conn.execute("""
        CREATE TABLE IF NOT EXISTS islem_gunlugu (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            personel_id INTEGER,
            personel_ad TEXT NOT NULL,
            islem_tipi TEXT NOT NULL,       -- "GİRİŞ", "OKUMA", "SİLME", "SYNC"
            detay TEXT,
            ip_adresi TEXT,
            tarih TEXT NOT NULL
        )
    """)
    
    # 2. Ürün Sorunları (Fotoğraflı hasar ve tutanak kayıtları)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS urun_sorunlari (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            uid TEXT NOT NULL,
            kategori TEXT NOT NULL,
            aciklama TEXT,
            ciddiyet TEXT,
            foto_yolu TEXT,
            tutanak_yolu TEXT,
            ekleyen_personel TEXT,
            tarih TEXT NOT NULL
        )
    """)

    # ── Migrasyonlar ──
    try:
        conn.execute("ALTER TABLE urunler ADD COLUMN parent_uid TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE birimler ADD COLUMN parent_uid TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE personeller ADD COLUMN sifre TEXT")
    except Exception:
        pass
        
    # Online sync kolonları ekle
    col_queries = [
        "ALTER TABLE urunler ADD COLUMN ekleyen_personel TEXT",
        "ALTER TABLE urunler ADD COLUMN silen_personel TEXT",
        "ALTER TABLE urunler ADD COLUMN sorun_foto_yolu TEXT",
        "ALTER TABLE urunler ADD COLUMN tutanak_belge_yolu TEXT"
    ]
    for q in col_queries:
        try:
            conn.execute(q)
        except Exception:
            pass

    # ── İndeksler ──
    conn.execute("CREATE INDEX IF NOT EXISTS idx_urunler_yukleme ON urunler(yukleme_id)")
    
    # Mükerrer temizliği (Başka bilgisayarlara taşınan eski veritabanları için)
    try:
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_urunler_uid_unique ON urunler(uid) WHERE deleted_at IS NULL")
    except Exception as e:
        if "UNIQUE constraint failed" in str(e):
            # Veritabanında zaten mükerrer kayıt var. Temizleyip tekrar deneyelim.
            conn.execute("""
                DELETE FROM urunler 
                WHERE id NOT IN (
                    SELECT MAX(id) 
                    FROM urunler 
                    WHERE deleted_at IS NULL 
                    GROUP BY uid
                ) 
                AND deleted_at IS NULL
            """)
            conn.commit()
            # Temizlik sonrası tekrar dene
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_urunler_uid_unique ON urunler(uid) WHERE deleted_at IS NULL")
            
    conn.execute("CREATE INDEX IF NOT EXISTS idx_urunler_sid ON urunler(sid)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cesitler_yukleme ON urun_cesitleri(yukleme_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_yuklemeler_depo ON stok_yuklemeler(depo_id)")

    conn.commit()
    return conn


def _stok_kaydet(params):
    """
    Stok kayıt işlemi — tam kapsamlı:
    1. XML+TXT dosyalarını yedek dizinine kopyala
    2. stok_yuklemeler tablosuna ana kayıt
    3. urun_cesitleri tablosuna SummaryItem'lar
    4. urunler tablosuna tekil ürünler (UID, ağırlık, SID, lot, SKT)
    """
    depo_id = params.get("depo_id")
    depo_adi = params.get("depo_adi", "")
    xml_yolu = params.get("xml_yolu", "")
    txt_yolu = params.get("txt_yolu", "")

    try:
        conn = _db_baglanti()
        now = datetime.now().isoformat()

        # ══════════════════════════════════════════════════════════════════
        # 1. DOSYA YEDEKLEME
        # ══════════════════════════════════════════════════════════════════
        yedek_dir = _yedek_dizini()
        zaman_ek = datetime.now().strftime("%Y%m%d_%H%M%S")
        xml_yedek = ""
        txt_yedek = ""

        if xml_yolu and os.path.exists(xml_yolu):
            xml_isim = os.path.basename(xml_yolu)
            xml_yedek = os.path.join(yedek_dir, f"{zaman_ek}_{xml_isim}")
            shutil.copy2(xml_yolu, xml_yedek)

        if txt_yolu and os.path.exists(txt_yolu):
            txt_isim = os.path.basename(txt_yolu)
            txt_yedek = os.path.join(yedek_dir, f"{zaman_ek}_{txt_isim}")
            shutil.copy2(txt_yolu, txt_yedek)

        # ══════════════════════════════════════════════════════════════════
        # 2. XML PARSE — gönderi bilgi + SummaryItem'lar
        # ══════════════════════════════════════════════════════════════════
        gonderi = {}
        summary_items = []
        sid_meta = {}  # SID → {urun_adi, lot, skt}

        if xml_yolu and os.path.exists(xml_yolu):
            parse_sonuc = _xml_parse(xml_yolu)
            if parse_sonuc.get("basarili"):
                gonderi = parse_sonuc.get("gonderi_bilgi", {})
                summary_items = parse_sonuc.get("summary_items", [])
                for si in summary_items:
                    sid_meta[si["sid"]] = {
                        "urun_adi": si.get("urun_adi", ""),
                        "lot": si.get("lot", ""),
                        "skt": si.get("skt", ""),
                    }

        # ══════════════════════════════════════════════════════════════════
        # 3. TXT PARSE — tekil ürünler
        # ══════════════════════════════════════════════════════════════════
        urunler = []
        if txt_yolu and os.path.exists(txt_yolu):
            txt_sonuc = _txt_parse(txt_yolu, xml_yolu)
            if txt_sonuc.get("basarili"):
                urunler = txt_sonuc.get("urunler", [])

        if not urunler:
            return {"basarili": False, "mesaj": "❌ Dosyada hiç ürün bulunamadı."}

        # ── Mükerrer kontrolü — Erken kontrol ve Dosya İçi Temizlik ──
        gelen_uids = [u["uid"] for u in urunler]
        placeholders = ",".join(["?"] * len(gelen_uids))
        mevcutlar = conn.execute(
            f"SELECT uid FROM urunler WHERE uid IN ({placeholders}) AND deleted_at IS NULL", 
            gelen_uids
        ).fetchall()
        mevcut_uid_set = {r[0] for r in mevcutlar}
        
        # Hem mevcut olanları atla, hem de dosya içindeki kendi mükerrerlerini temizle
        unique_urunler = []
        seen = set()
        for u in urunler:
            uid = u["uid"]
            if uid not in mevcut_uid_set and uid not in seen:
                seen.add(uid)
                unique_urunler.append(u)
                
        mukerrer_sayisi = len(urunler) - len(unique_urunler)
        urunler = unique_urunler
        
        if not urunler:
            return {"basarili": False, "mesaj": "⚠️ Bu dosyadaki tüm ürünler zaten stokta mevcut (veya kendi içinde mükerrer). Tekrar yükleme yapılamaz."}

        # ══════════════════════════════════════════════════════════════════
        # 4. ANA KAYIT — stok_yuklemeler
        # ══════════════════════════════════════════════════════════════════
        toplam_agirlik = sum(float(u.get("agirlik", 0)) for u in urunler)

        cursor = conn.execute(
            """INSERT INTO stok_yuklemeler 
               (depo_id, depo_adi, gonderi_no, mesaj_id, gonderen, alici, 
                teslim_tarihi, xml_dosya, txt_dosya, xml_yedek, txt_yedek,
                urun_sayisi, cesit_sayisi, toplam_agirlik, yukleme_tarihi) 
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                depo_id, depo_adi,
                gonderi.get("gonderi_no", ""),
                gonderi.get("mesaj_id", ""),
                gonderi.get("gonderen", ""),
                gonderi.get("alici", ""),
                gonderi.get("teslim_tarihi", ""),
                xml_yolu, txt_yolu,
                xml_yedek, txt_yedek,
                len(urunler),
                len(summary_items),
                round(toplam_agirlik, 3),
                now,
            ),
        )
        yukleme_id = cursor.lastrowid

        # ══════════════════════════════════════════════════════════════════
        # 5. ÜRÜN ÇEŞİTLERİ — urun_cesitleri
        # ══════════════════════════════════════════════════════════════════
        for si in summary_items:
            conn.execute(
                """INSERT INTO urun_cesitleri
                   (yukleme_id, sid, psn, urun_kodu, urun_adi, lot, skt, uretim_tarihi, adet)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    yukleme_id, si.get("sid"), si.get("psn"),
                    si.get("urun_kodu"), si.get("urun_adi"),
                    si.get("lot"), si.get("skt"), si.get("uretim_tarihi"),
                    si.get("adet", 0),
                ),
            )

        # ══════════════════════════════════════════════════════════════════
        # 6. TEKİL ÜRÜNLER — urunler (batch insert)
        # ══════════════════════════════════════════════════════════════════
        parent_map = parse_sonuc.get("parent_map", {}) if "parse_sonuc" in locals() else {}
        
        if urunler:
            conn.executemany(
                """INSERT OR IGNORE INTO urunler
                   (yukleme_id, uid, psn, sid, kod, paketleme, agirlik, 
                    xml_eslesme, urun_adi, lot, skt, parent_uid)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                [
                    (
                        yukleme_id,
                        u["uid"], u.get("psn"), u.get("sid"), u.get("kod"),
                        u.get("paketleme"),
                        float(u.get("agirlik", 0)),
                        1 if u.get("xml_eslesme") else 0,
                        sid_meta.get(u.get("sid"), {}).get("urun_adi", ""),
                        sid_meta.get(u.get("sid"), {}).get("lot", ""),
                        sid_meta.get(u.get("sid"), {}).get("skt", ""),
                        parent_map.get(u["uid"])
                    )
                    for u in urunler
                ],
            )

        # ══════════════════════════════════════════════════════════════════
        # 7. BİRİMLERİ KAYDET (Koli/Kutu Meta)
        # ══════════════════════════════════════════════════════════════════
        unit_details = parse_sonuc.get("unit_details", []) if "parse_sonuc" in locals() else []
        if unit_details:
            conn.executemany(
                "INSERT OR IGNORE INTO birimler (yukleme_id, uid, seviye, adet, parent_uid) VALUES (?, ?, ?, ?, ?)",
                [(yukleme_id, ud["uid"], ud["seviye"], ud["adet"], ud.get("parent_uid")) for ud in unit_details]
            )

        conn.commit()
        conn.close()

        # ── Özet bilgi ──
        yedek_bilgi = ""
        if xml_yedek or txt_yedek:
            yedek_bilgi = f" Dosyalar yedeklendi."

        # ── Özet bilgi ──
        durum_mesaj = f"✅ {len(urunler)} yeni ürün"
        if mukerrer_sayisi > 0:
            durum_mesaj += f" ({mukerrer_sayisi} ürün zaten stoktaydı, atlandı)"
        
        durum_mesaj += f" \"{depo_adi}\" deposuna kaydedildi."

        return {
            "basarili": True,
            "yukleme_id": yukleme_id,
            "mesaj": durum_mesaj,
            "detay": {
                "yukleme_id": yukleme_id,
                "urun_sayisi": len(urunler),
                "mukerrer_sayisi": mukerrer_sayisi,
                "cesit_sayisi": len(summary_items),
                "toplam_agirlik": round(toplam_agirlik, 3),
                "xml_yedek": xml_yedek,
                "txt_yedek": txt_yedek,
                "db_yolu": _db_yolu(),
            },
        }

    except Exception as e:
        return {
            "basarili": False,
            "mesaj": f"❌ Kayıt hatası: {e}",
        }


# ══════════════════════════════════════════════════════════════════════════════
# STOK ÖZET — Genel Envanter Görünümü
# ══════════════════════════════════════════════════════════════════════════════

def _stok_ozet(params):
    """
    Tüm depoların ve ürünlerin özetini getirir.
    """
    depo_id = params.get("depo_id")
    
    try:
        conn = _db_baglanti()
        
        # 1. Genel İstatistikler
        stats_query = """
            SELECT 
                COUNT(*) as toplam_urun,
                SUM(agirlik) as toplam_agirlik,
                COUNT(DISTINCT sid) as cesit_sayisi,
                COUNT(DISTINCT parent_uid) as kutu_sayisi
            FROM urunler 
            WHERE deleted_at IS NULL
        """
        if depo_id:
            stats_query += " AND yukleme_id IN (SELECT id FROM stok_yuklemeler WHERE depo_id = ?)"
            stats = conn.execute(stats_query, (depo_id,)).fetchone()
        else:
            stats = conn.execute(stats_query).fetchone()

        # 2. Depo Bazlı Dağılım
        depolar = conn.execute("""
            SELECT 
                d.id, d.kod, d.ad,
                COUNT(u.id) as urun_sayisi,
                ROUND(SUM(u.agirlik), 2) as agirlik
            FROM depolar d
            LEFT JOIN stok_yuklemeler y ON y.depo_id = d.id
            LEFT JOIN urunler u ON u.yukleme_id = y.id AND u.deleted_at IS NULL
            GROUP BY d.id
        """).fetchall()

        # 3. Ürün Çeşidi Bazlı Dağılım (Top 20)
        cesit_query = """
            SELECT 
                sid, urun_adi,
                COUNT(*) as adet,
                ROUND(SUM(agirlik), 2) as agirlik
            FROM urunler
            WHERE deleted_at IS NULL
        """
        if depo_id:
            cesit_query += " AND yukleme_id IN (SELECT id FROM stok_yuklemeler WHERE depo_id = ?)"
            cesit_query += " GROUP BY sid ORDER BY adet DESC LIMIT 20"
            cesitler = conn.execute(cesit_query, (depo_id,)).fetchall()
        else:
            cesit_query += " GROUP BY sid ORDER BY adet DESC LIMIT 20"
            cesitler = conn.execute(cesit_query).fetchall()

        # 4. Son Yüklemeler
        yukleme_query = """
            SELECT id, depo_adi, gonderi_no, urun_sayisi, yukleme_tarihi
            FROM stok_yuklemeler
            WHERE deleted_at IS NULL
        """
        if depo_id:
            yukleme_query += " AND depo_id = ?"
            yukleme_query += " ORDER BY yukleme_tarihi DESC LIMIT 5"
            yuklemeler = conn.execute(yukleme_query, (depo_id,)).fetchall()
        else:
            yukleme_query += " ORDER BY yukleme_tarihi DESC LIMIT 5"
            yuklemeler = conn.execute(yukleme_query).fetchall()

        conn.close()
        
        return {
            "basarili": True,
            "stats": dict(stats) if stats else {},
            "depolar": [dict(r) for r in depolar],
            "cesitler": [dict(r) for r in cesitler],
            "son_yuklemeler": [dict(r) for r in yuklemeler]
        }
    except Exception as e:
        return {"basarili": False, "mesaj": f"Stok özeti alınamadı: {e}"}


# ══════════════════════════════════════════════════════════════════════════════
# STOK YÖNETİMİ (Sıfırlama, Çıkış, Envanter)
# ══════════════════════════════════════════════════════════════════════════════

def _stok_sifirla(params):
    """Stok verilerini gizler veya tamamen siler."""
    mod = params.get("mod", "soft")
    try:
        conn = _db_baglanti()
        now = datetime.now().isoformat()
        
        if mod == "soft":
            # Yumuşak silme: deleted_at işaretle
            conn.execute("UPDATE stok_yuklemeler SET deleted_at = ? WHERE deleted_at IS NULL", (now,))
            conn.execute("UPDATE urunler SET deleted_at = ? WHERE deleted_at IS NULL", (now,))
            mesaj = "Tüm stok verileri başarıyla gizlendi (Arşive alındı)."
        else:
            # Sert silme: Tabloları boşalt
            conn.execute("DELETE FROM urunler")
            conn.execute("DELETE FROM urun_cesitleri")
            conn.execute("DELETE FROM stok_yuklemeler")
            conn.execute("DELETE FROM birimler")
            mesaj = "Tüm stok veritabanı tabloları başarıyla boşaltıldı."
            
        conn.commit()
        conn.close()
        return {"basarili": True, "mesaj": mesaj}
    except Exception as e:
        return {"basarili": False, "mesaj": f"Sıfırlama hatası: {e}"}


def _stok_veritabani_sil():
    """Veritabanı dosyasını tamamen siler."""
    try:
        db_path = _db_yolu()
        if os.path.exists(db_path):
            # Dosya silmeden önce WAL ve SHM dosyalarını da temizlemeye çalışalım
            for ext in ["", "-wal", "-shm"]:
                p = db_path + ext
                if os.path.exists(p):
                    try:
                        os.remove(p)
                    except:
                        pass
        return {"basarili": True, "mesaj": "Veritabanı dosyası başarıyla silindi. Sistem yeniden başlatılıyor."}
    except Exception as e:
        return {"basarili": False, "mesaj": f"Dosya silme hatası: {e}"}


def _stok_cikis_onizle(params):
    """
    UID listesi verilen ürünlerin durumunu kontrol eder. 
    Hiyerarşik (Koli/Palet) UID'leri tekil ürünlere açar.
    """
    input_uids = params.get("uid_listesi", [])
    if not input_uids:
        return {"basarili": False, "mesaj": "UID listesi boş."}
    
    conn = None
    try:
        conn = _db_baglanti()
        
        uid_to_raw = {}
        cleaned_uids = []
        for raw in input_uids:
            raw = raw.strip()
            if not raw: continue
            # 1. (250)UID formatı varsa ayıkla
            m = re.search(r"\(250\)([^(]+)", raw)
            if m:
                uid = m.group(1).strip()
            # 2. 51 haneli parantezsiz format (90...250UID240...)
            elif len(raw) >= 42 and raw.startswith("90") and "250" in raw and "240" in raw:
                m2 = re.search(r"^90.{6}250(.{11})240", raw)
                uid = m2.group(1).strip() if m2 else raw
            else:
                uid = raw
            cleaned_uids.append(uid)
            uid_to_raw[uid] = raw
            
        final_items = {} # uid -> item_dict
        bulunamayan = []
        bulunamayan_detayli = []
        
        for uid in cleaned_uids:
            # A. Önce doğrudan ürün mü diye bak
            item = conn.execute("""
                SELECT u.uid, u.urun_adi, u.psn, u.kod, u.agirlik, u.lot, u.skt, u.yukleme_id, u.parent_uid,
                       y.depo_id, d.kod as depo_kod, d.ad as depo_adi
                FROM urunler u
                LEFT JOIN stok_yuklemeler y ON u.yukleme_id = y.id
                LEFT JOIN depolar d ON y.depo_id = d.id
                WHERE u.uid = ? AND u.deleted_at IS NULL
            """, (uid,)).fetchone()
            
            if item:
                d_item = dict(item)
                d_item["raw_gs1"] = uid_to_raw.get(uid, uid)
                final_items[item["uid"]] = d_item
                continue
                
            # B. Değilse bir 'Birim' (Koli/Palet) mi diye bak
            child_items = conn.execute("""
                SELECT u.uid, u.urun_adi, u.psn, u.kod, u.agirlik, u.lot, u.skt, u.yukleme_id, u.parent_uid,
                       y.depo_id, d.kod as depo_kod, d.ad as depo_adi
                FROM urunler u
                LEFT JOIN stok_yuklemeler y ON u.yukleme_id = y.id
                LEFT JOIN depolar d ON y.depo_id = d.id
                WHERE (u.parent_uid = ? OR u.parent_uid IN (SELECT uid FROM birimler WHERE parent_uid = ?))
                AND u.deleted_at IS NULL
            """, (uid, uid)).fetchall()
            
            if child_items:
                for ci in child_items:
                    d_ci = dict(ci)
                    # Koli okutulunca içindeki tekillerin raw GS1'i elimizde yok, kendimiz üretmeliyiz
                    # Ama şimdilik sadece UID saklayalım, stok_cikis'te halledeceğiz
                    d_ci["raw_gs1"] = d_ci["uid"]
                    final_items[ci["uid"]] = d_ci
            else:
                raw_gs1 = uid_to_raw.get(uid, uid)
                m_psn = re.search(r"\(90\)([^(]+)", raw_gs1)
                m_kod = re.search(r"\(240\)([^(]+)", raw_gs1)
                m_agirlik = re.search(r"\(3[01]\d{2}\)(\d+)", raw_gs1)
                m_lot = re.search(r"\(10\)([^(]+)", raw_gs1)
                m_skt = re.search(r"\(17\)([^(]+)", raw_gs1)
                
                if m_psn and m_kod:
                    agirlik_kg = float(m_agirlik.group(1))/1000 if m_agirlik else 0
                    lot_val = m_lot.group(1).strip() if m_lot else ""
                    skt_val = m_skt.group(1).strip() if m_skt else ""
                    bulunamayan_detayli.append({
                        "uid": uid,
                        "raw_gs1": raw_gs1,
                        "psn": m_psn.group(1).strip(),
                        "kod": m_kod.group(1).strip(),
                        "agirlik": agirlik_kg,
                        "urun_adi": "Depoda Yok (GS1)",
                        "lot": lot_val,
                        "skt": skt_val,
                        "parse_edildi": True
                    })
                else:
                    # Sadece UID varsa (GS1 formatı değilse) yine manuel eklenebilir yap
                    bulunamayan_detayli.append({
                        "uid": uid,
                        "raw_gs1": raw_gs1,
                        "psn": "BİLİNMEYEN",
                        "kod": "000",
                        "agirlik": 0.0,
                        "urun_adi": "Depoda Yok (Sadece UID)",
                        "lot": "",
                        "skt": "",
                        "parse_edildi": False
                    })
                
        bulunanlar = list(final_items.values())
        toplam_agirlik = sum(float(r["agirlik"] or 0) for r in bulunanlar)
        
        return {
            "basarili": True,
            "urunler": bulunanlar,
            "bulunamayan": bulunamayan,
            "bulunamayan_detayli": bulunamayan_detayli,
            "ozet": {
                "taranan": len(cleaned_uids),
                "acilan_urun": len(bulunanlar),
                "bulunamayan": len(bulunamayan) + len(bulunamayan_detayli),
                "toplam_agirlik": round(toplam_agirlik, 3)
            }
        }
    except Exception as e:
        return {"basarili": False, "mesaj": f"Önizleme hatası: {e}"}
    finally:
        if conn:
            conn.close()


def _stok_cikis(params):
    """Onaylanan ürünleri stoktan düşer."""
    uids = params.get("uid_listesi", [])
    raw_gs1_map = params.get("raw_gs1_map", {})
    manuel = params.get("manuel_eklenenler", [])
    depo_id = params.get("depo_id")
    sebep = params.get("sebep", "Çıkış")
    
    if not uids and not manuel:
        return {"basarili": False, "mesaj": "Çıkış listesi boş."}
    
    conn = None
    try:
        conn = _db_baglanti()
        now = datetime.now().isoformat()
        
        # GS1 listesi oluştur (Frontend'in TXT indirmesi için)
        gs1_listesi = []
        toplam_agirlik = 0
        urun_sayisi = 0
        
        if uids:
            placeholders = ",".join(["?"] * len(uids))
            rows = conn.execute(f"SELECT * FROM urunler WHERE uid IN ({placeholders}) AND deleted_at IS NULL", uids).fetchall()
            
            # 2. Silme (soft-delete)
            conn.execute(f"UPDATE urunler SET deleted_at = ? WHERE uid IN ({placeholders})", [now] + uids)
            
            for r in rows:
                uid_val = r["uid"]
                if uid_val in raw_gs1_map and raw_gs1_map[uid_val]:
                    gs1_listesi.append(raw_gs1_map[uid_val])
                else:
                    ag = float(r["agirlik"] or 0)
                    agirlik_ham = str(int(ag * 1000)).zfill(6)
                    gs1 = f"(90){r['psn']}(250){r['uid']}(240){r['kod']}(20)00(3103){agirlik_ham}"
                    gs1_listesi.append(gs1)
                    
                toplam_agirlik += float(r["agirlik"] or 0)
                urun_sayisi += 1
                
        if manuel:
            # Manuel eklenenler için sahte bir yükleme kaydı (log için)
            c = conn.cursor()
            c.execute("""
                INSERT INTO stok_yuklemeler (depo_id, gonderi_no, gonderen, alici, yukleme_tarihi)
                VALUES (?, ?, ?, ?, ?)
            """, (depo_id, "MANUEL ÇIKIŞ", "Sistem", sebep, now))
            yukleme_id = c.lastrowid
            
            for m in manuel:
                ag = float(m.get("agirlik", 0))
                # DB'ye hemen "silinmiş" olarak ekle ki Hareketler raporunda görünsün
                c.execute("""
                    INSERT INTO urunler (uid, psn, kod, urun_adi, agirlik, lot, skt, yukleme_id, deleted_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (m["uid"], m.get("psn", ""), m.get("kod", ""), m.get("urun_adi", ""), 
                      ag, m.get("lot", ""), m.get("skt", ""), yukleme_id, now))
                
                # Orijinal GS1 metni varsa direkt onu kullan, yoksa kendin üret
                raw = m.get("raw_gs1")
                if raw:
                    gs1_listesi.append(raw)
                else:
                    agirlik_ham = str(int(ag * 1000)).zfill(6)
                    gs1 = f"(90){m.get('psn','')}(250){m['uid']}(240){m.get('kod','')}(20)00(3103){agirlik_ham}"
                    gs1_listesi.append(gs1)
                    
                toplam_agirlik += ag
                urun_sayisi += 1
            
        cikis_no = f"OUT-{datetime.now().strftime('%y%m%d%H%M')}"
        
        conn.commit()
        
        return {
            "basarili": True,
            "mesaj": f"✅ {urun_sayisi} ürün başarıyla çıkış yapıldı.",
            "cikis_no": cikis_no,
            "urun_sayisi": urun_sayisi,
            "toplam_agirlik": round(toplam_agirlik, 3),
            "gs1_listesi": gs1_listesi
        }
    except Exception as e:
        if conn:
            conn.rollback()
        return {"basarili": False, "mesaj": f"Çıkış hatası: {e}"}
    finally:
        if conn:
            conn.close()


def _stok_envanter(params):
    """Gelişmiş envanter sorgulama ve filtreleme."""
    depo_id = params.get("depo_id")
    arama = params.get("arama", "").strip()
    grup = params.get("grup", "hepsi")
    sayfa = params.get("sayfa", 1)
    limit = params.get("sayfa_boyut", 100)
    offset = (sayfa - 1) * limit
    
    try:
        conn = _db_baglanti()
        
        # Temel sorgu
        if grup == "cesit":
            base_query = """
                SELECT urun_adi, sid, psn, kod as urun_kodu, COUNT(*) as adet, SUM(agirlik) as toplam_agirlik
                FROM urunler
                WHERE deleted_at IS NULL
            """
            group_by = "GROUP BY sid"
        elif grup == "kutu":
            base_query = """
                SELECT parent_uid as uid, urun_adi, COUNT(*) as adet, SUM(agirlik) as toplam_agirlik
                FROM urunler
                WHERE deleted_at IS NULL AND parent_uid IS NOT NULL
            """
            group_by = "GROUP BY parent_uid"
        else:
            base_query = """
                SELECT * FROM urunler
                WHERE deleted_at IS NULL
            """
            group_by = ""
            
        where_clauses = []
        query_params = []
        
        if depo_id:
            where_clauses.append("yukleme_id IN (SELECT id FROM stok_yuklemeler WHERE depo_id = ?)")
            query_params.append(depo_id)
            
        if arama:
            where_clauses.append("(urun_adi LIKE ? OR uid LIKE ? OR sid LIKE ?)")
            p = f"%{arama}%"
            query_params.extend([p, p, p])
            
        if where_clauses:
            if "WHERE" in base_query:
                base_query += " AND " + " AND ".join(where_clauses)
            else:
                base_query += " WHERE " + " AND ".join(where_clauses)
                
        # Toplam sayısı (pagination için)
        count_query = f"SELECT COUNT(*) FROM ({base_query} {group_by})"
        toplam_row = conn.execute(count_query, query_params).fetchone()
        toplam = toplam_row[0] if toplam_row else 0
        
        # Veri sorgusu
        data_query = f"{base_query} {group_by} ORDER BY id DESC LIMIT ? OFFSET ?"
        rows = conn.execute(data_query, query_params + [limit, offset]).fetchall()
        
        conn.close()
        return {
            "basarili": True,
            "data": [dict(r) for r in rows],
            "toplam": toplam,
            "sayfa": sayfa,
            "limit": limit
        }
    except Exception as e:
        return {"basarili": False, "mesaj": f"Envanter hatası: {e}"}


def _stok_sayim_karsilastir(params):
    """Okutulan UID listesini stokla karşılaştırır."""
    uid_listesi = params.get("uid_listesi", [])
    depo_id = params.get("depo_id")
    
    if not uid_listesi:
        return {"basarili": False, "mesaj": "Okutulan barkod yok."}
    
    try:
        conn = _db_baglanti()
        
        # 1. Depodaki tüm aktif ürünleri getir
        query = "SELECT uid FROM urunler WHERE deleted_at IS NULL"
        q_params = []
        if depo_id:
            query += " AND yukleme_id IN (SELECT id FROM stok_yuklemeler WHERE depo_id = ?)"
            q_params.append(depo_id)
            
        db_uids = {r["uid"] for r in conn.execute(query, q_params).fetchall()}
        scan_uids = set(uid_listesi)
        
        eslesen = scan_uids.intersection(db_uids)
        fazla = scan_uids.difference(db_uids)
        eksik = db_uids.difference(scan_uids)
        
        # Eksik ve fazla olanların detaylarını getir (ilk 100 tanesini)
        def get_details(uids_set):
            if not uids_set: return []
            uids_list = list(uids_set)[:100]
            ph = ",".join(["?"] * len(uids_list))
            rows_detay = conn.execute(f"SELECT uid, urun_adi, psn FROM urunler WHERE uid IN ({ph})", uids_list).fetchall()
            return [dict(r) for r in rows_detay]
            
        fazla_detay = get_details(fazla)
        eksik_detay = get_details(eksik)

        conn.close()
        return {
            "basarili": True,
            "ozet": {
                "eslesen_adet": len(eslesen),
                "fazla_adet": len(fazla),
                "eksik_adet": len(eksik),
                "toplam_okutulan": len(scan_uids),
                "toplam_stok": len(db_uids)
            },
            "eslesen": list(eslesen)[:100],
            "fazla_detay": fazla_detay,
            "eksik_detay": eksik_detay
        }
    except Exception as e:
        return {"basarili": False, "mesaj": f"Karşılaştırma hatası: {e}"}


# ══════════════════════════════════════════════════════════════
# RAPOR VERİSİ — UI dashboard için (Veri Doğrulanmış)
# ══════════════════════════════════════════════════════════════

def _rapor_veri(params):
    """
    Dashboard ve önizleme tabloları için veri çeker.
    """
    depo_id   = params.get("depo_id")
    baslangic = params.get("baslangic")
    bitis     = params.get("bitis")

    try:
        conn = _db_baglanti()

        depo_where = "AND y.depo_id = ?" if depo_id else ""
        depo_bind  = [depo_id] if depo_id else []

        tarih_w, tarih_b = "", []
        if baslangic:
            tarih_w += " AND y.yukleme_tarihi >= ?"; tarih_b.append(baslangic)
        if bitis:
            tarih_w += " AND y.yukleme_tarihi <= ?"; tarih_b.append(bitis + "T23:59:59")

        # ── Özet ──
        ozet_row = conn.execute(f"""
            SELECT
                COUNT(DISTINCT u.uid) as toplam_giris,
                COUNT(DISTINCT CASE WHEN u.deleted_at IS NULL THEN u.uid END) as net_stok,
                ROUND(SUM(CASE WHEN u.deleted_at IS NULL THEN u.agirlik ELSE 0 END), 2) as net_agirlik,
                ROUND(SUM(u.agirlik), 2) as giris_agirlik
            FROM urunler u
            JOIN stok_yuklemeler y ON y.id = u.yukleme_id
            WHERE 1=1 {depo_where} {tarih_w}
        """, depo_bind + tarih_b).fetchone()

        try:
            cw = "AND depo_id = ?" if depo_id else ""
            cb = [depo_id] if depo_id else []
            cikis_r = conn.execute(
                f"SELECT COALESCE(SUM(urun_sayisi),0) as adet, COALESCE(SUM(toplam_agirlik),0) as ag FROM cikislar WHERE deleted_at IS NULL {cw}", cb
            ).fetchone()
            toplam_cikis  = int(cikis_r["adet"])
            cikis_agirlik = float(cikis_r["ag"])
        except:
            toplam_cikis = 0; cikis_agirlik = 0.0

        ozet = {
            "toplam_giris":  int(ozet_row["toplam_giris"]  or 0),
            "giris_agirlik": float(ozet_row["giris_agirlik"] or 0),
            "net_stok":      int(ozet_row["net_stok"]      or 0),
            "net_agirlik":   float(ozet_row["net_agirlik"]  or 0),
            "toplam_cikis":  toplam_cikis,
            "cikis_agirlik": cikis_agirlik,
        }

        # ── Depo dağılımı ──
        depolar = conn.execute("SELECT id, kod, ad FROM depolar WHERE deleted_at IS NULL ORDER BY ad").fetchall()
        depo_dagilim = []
        for d in depolar:
            if depo_id and d["id"] != depo_id: continue
            s = conn.execute(f"""
                SELECT COUNT(DISTINCT u.uid) as adet,
                       ROUND(SUM(CASE WHEN u.deleted_at IS NULL THEN u.agirlik ELSE 0 END), 2) as agirlik
                FROM urunler u
                JOIN stok_yuklemeler y ON y.id = u.yukleme_id
                WHERE y.depo_id = ? AND u.deleted_at IS NULL {tarih_w}
            """, [d["id"]] + tarih_b).fetchone()
            try:
                c2 = conn.execute(
                    "SELECT COALESCE(SUM(urun_sayisi),0) as adet, COALESCE(SUM(toplam_agirlik),0) as ag FROM cikislar WHERE depo_id=? AND deleted_at IS NULL", (d["id"],)
                ).fetchone()
                c_adet = int(c2["adet"]); c_ag = float(c2["ag"])
            except:
                c_adet = 0; c_ag = 0.0
            depo_dagilim.append({
                "depo_kod": d["kod"], "depo_adi": d["ad"],
                "stok_adet": int(s["adet"] or 0),
                "cikis_adet": c_adet,
                "net_agirlik": float(s["agirlik"] or 0),
                "cikis_agirlik": c_ag,
            })

        # ── Son girişler ──
        son_girisler = conn.execute(f"""
            SELECT SUBSTR(y.yukleme_tarihi,1,16) as tarih,
                   d.ad as depo_adi, y.gonderi_no, y.gonderen,
                   y.urun_sayisi, y.cesit_sayisi, y.toplam_agirlik
            FROM stok_yuklemeler y
            LEFT JOIN depolar d ON d.id = y.depo_id
            WHERE y.deleted_at IS NULL {depo_where} {tarih_w}
            ORDER BY y.yukleme_tarihi DESC LIMIT 20
        """, depo_bind + tarih_b).fetchall()

        # ── Son çıkışlar ──
        try:
            cw2 = "AND c.depo_id = ?" if depo_id else ""
            cb2 = [depo_id] if depo_id else []
            son_cikislar = conn.execute(f"""
                SELECT SUBSTR(c.islem_tarihi,1,16) as tarih,
                       d.ad as depo_adi, c.cikis_no, c.sebep,
                       c.urun_sayisi, c.toplam_agirlik
                FROM cikislar c
                LEFT JOIN depolar d ON d.id = c.depo_id
                WHERE c.deleted_at IS NULL {cw2}
                ORDER BY c.islem_tarihi DESC LIMIT 20
            """, cb2).fetchall()
            son_cikislar = [dict(r) for r in son_cikislar]
        except:
            son_cikislar = []

        # ── Lot dağılımı ──
        lot_dagilim = conn.execute(f"""
            SELECT u.lot, u.urun_adi, u.skt, u.sid,
                   COUNT(DISTINCT u.uid) as adet,
                   ROUND(SUM(u.agirlik), 2) as agirlik,
                   d.ad as depo_adi, d.kod as depo_kod
            FROM urunler u
            JOIN stok_yuklemeler y ON y.id = u.yukleme_id
            LEFT JOIN depolar d ON d.id = y.depo_id
            WHERE u.deleted_at IS NULL {depo_where} {tarih_w}
            GROUP BY u.lot, u.sid, d.id
            ORDER BY adet DESC LIMIT 100
        """, depo_bind + tarih_b).fetchall()

        conn.close()
        return {
            "basarili": True,
            "ozet": ozet,
            "depo_dagilim": depo_dagilim,
            "son_girisler": [dict(r) for r in son_girisler],
            "son_cikislar": son_cikislar,
            "lot_dagilim":  [dict(r) for r in lot_dagilim],
        }
    except Exception as e:
        return {"basarili": False, "mesaj": f"Rapor veri hatası: {e}"}


def _rapor_olustur(params):
    """
    Excel raporu üretir.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"basarili": False, "mesaj": "openpyxl kurulu değil: pip install openpyxl"}

    depo_id   = params.get("depo_id")
    baslangic = params.get("baslangic", "")
    bitis     = params.get("bitis", "")

    veri = _rapor_veri(params)
    if not veri.get("basarili"):
        return veri

    # ── Ürün listesini ayrıca çek (unique uid) ──
    try:
        conn = _db_baglanti()
        depo_w = "AND y.depo_id = ?" if depo_id else ""
        db_b   = [depo_id] if depo_id else []
        tw, tb = "", []
        if baslangic: tw += " AND y.yukleme_tarihi >= ?"; tb.append(baslangic)
        if bitis:     tw += " AND y.yukleme_tarihi <= ?"; tb.append(bitis+"T23:59:59")

        urun_rows = conn.execute(f"""
            SELECT u.uid, u.psn, u.sid, u.kod, u.urun_adi, u.lot, u.skt,
                   u.agirlik, u.deleted_at,
                   d.ad as depo_adi, d.kod as depo_kod, y.gonderi_no
            FROM urunler u
            JOIN stok_yuklemeler y ON y.id = u.yukleme_id
            LEFT JOIN depolar d ON d.id = y.depo_id
            WHERE 1=1 {depo_w} {tw}
            ORDER BY d.ad, u.urun_adi, u.uid
        """, db_b + tb).fetchall()
        conn.close()

        # Distinct uid
        seen, urun_uniq = set(), []
        for u in urun_rows:
            if u["uid"] not in seen:
                seen.add(u["uid"])
                urun_uniq.append(dict(u))
    except Exception as e:
        urun_uniq = []

    # ── Stiller ──
    KOYU_MAVI  = "0D2137"; KOYU_YESIL = "1A3C34"; TURUNCU = "E65100"
    KIRMIZI = "B71C1C"; GRI = "546E7A"; BEYAZ = "FFFFFF"; SIYAH = "212121"
    ACIK_MAVI="E3F2FD"; ACIK_YESIL="E8F5E9"; ACIK_KIRM="FFEBEE"; ACIK_GRI="F5F5F5"

    def kenar():
        k = Side(style="thin", color="CCCCCC")
        return Border(left=k, right=k, top=k, bottom=k)

    def th(ws, row, col, val, bg=None, renk=None, boyut=9, bold=True, hizala="center"):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=bold, color=renk or BEYAZ, size=boyut)
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=hizala, vertical="center", wrap_text=True)
        c.border = kenar()
        return c

    def td(ws, row, col, val, bg=None, renk=None, sag=False, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", size=9, color=renk or SIYAH, bold=bold)
        c.alignment = Alignment(horizontal="right" if sag else "left", vertical="center")
        c.border = kenar()
        return c

    wb = openpyxl.Workbook()

    # ════ SEKME 1: ÖZET ════
    ws1 = wb.active; ws1.title = "Özet"
    ws1.sheet_view.showGridLines = False
    for col, w in [("A",30),("B",16),("C",16),("D",16),("E",16)]:
        ws1.column_dimensions[col].width = w

    ws1.merge_cells("A1:E1")
    c = ws1["A1"]
    c.value = "GAMALI PATBİS — STOK ENVANTER RAPORU"
    c.font = Font(bold=True, color=BEYAZ, size=15)
    c.fill = PatternFill("solid", fgColor=KOYU_MAVI)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells("A2:E2")
    c2 = ws1["A2"]
    depo_txt = "Tüm Depolar"
    if depo_id:
        try:
            conn2 = _db_baglanti()
            d = conn2.execute("SELECT ad FROM depolar WHERE id=?", (depo_id,)).fetchone()
            if d: depo_txt = d["ad"]
            conn2.close()
        except: pass
    c2.value = f"Oluşturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  {depo_txt}  |  {baslangic or 'Başlangıç'} — {bitis or 'Bugün'}"
    c2.font = Font(color=GRI, size=9, italic=True)
    c2.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 16

    # Özet kartları
    oz = veri["ozet"]
    kartlar = [
        ("Toplam Giriş",   oz["toplam_giris"],              KOYU_MAVI,  ACIK_MAVI),
        ("Net Stok",       oz["net_stok"],                  KOYU_YESIL, ACIK_YESIL),
        ("Toplam Çıkış",   oz["toplam_cikis"],              KIRMIZI,    ACIK_KIRM),
        ("Giriş Ağırlık",  f"{oz['giris_agirlik']} kg",    GRI,        ACIK_GRI),
        ("Net Ağırlık",    f"{oz['net_agirlik']} kg",       KOYU_YESIL, ACIK_YESIL),
    ]
    for i, (lbl, val, renk, bg) in enumerate(kartlar):
        col = i+1
        ws1.row_dimensions[4].height = 14
        ws1.row_dimensions[5].height = 30
        th(ws1, 4, col, lbl, bg=bg, renk=renk, boyut=8)
        c = ws1.cell(row=5, column=col, value=val)
        c.font = Font(bold=True, color=renk, size=15)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = kenar()

    # Depo tablosu
    ws1.row_dimensions[7].height = 18
    ws1.merge_cells("A7:E7")
    h = ws1["A7"]; h.value="DEPO BAZLI DAĞILIM"
    h.font=Font(bold=True,color=BEYAZ,size=11); h.fill=PatternFill("solid",fgColor=KOYU_MAVI)
    h.alignment=Alignment(horizontal="center")

    for j,b in enumerate(["Depo Adı","Depo Kodu","Net Stok","Net Ağırlık (kg)","Çıkış (adet)"]):
        th(ws1, 8, j+1, b, bg=GRI)
    ws1.row_dimensions[8].height = 16

    for ri, d in enumerate(veri["depo_dagilim"]):
        row = 9+ri; ws1.row_dimensions[row].height = 15
        bg = ACIK_GRI if ri%2==1 else BEYAZ
        td(ws1,row,1,d["depo_adi"],bg=bg)
        td(ws1,row,2,d["depo_kod"],bg=bg,renk=GRI)
        td(ws1,row,3,d["stok_adet"],bg=bg,renk=KOYU_YESIL,sag=True,bold=bool(d["stok_adet"]))
        td(ws1,row,4,d["net_agirlik"],bg=bg,sag=True)
        td(ws1,row,5,d["cikis_adet"],bg=bg,renk=KIRMIZI if d["cikis_adet"]>0 else GRI,sag=True)

    # ════ SEKME 2: ÜRÜN BAZLI ════
    ws2 = wb.create_sheet("Ürün Bazlı")
    ws2.sheet_view.showGridLines = False
    s2 = [("UID",20),("PSN",10),("SID",14),("Ürün Kodu",16),("Ürün Adı",32),
          ("Lot",16),("SKT",12),("Ağırlık (kg)",14),("Depo",26),("Depo Kodu",14),
          ("Gönderi No",16),("Durum",10)]
    for j,(b,w) in enumerate(s2):
        ws2.column_dimensions[get_column_letter(j+1)].width=w
        th(ws2,1,j+1,b,bg=KOYU_MAVI)
    ws2.row_dimensions[1].height=18; ws2.freeze_panes="A2"

    for ri, u in enumerate(urun_uniq):
        row=ri+2; ws2.row_dimensions[row].height=13
        durum = "Çıkış" if u["deleted_at"] else "Stokta"
        bg = "FFF3E0" if u["deleted_at"] else (ACIK_GRI if ri%2==1 else BEYAZ)
        vals=[u["uid"],u["psn"],u["sid"],u["kod"],u["urun_adi"],
              u["lot"],u["skt"],u["agirlik"],u["depo_adi"],u["depo_kod"],
              u["gonderi_no"],durum]
        for j,v in enumerate(vals):
            renk = KIRMIZI if (j==11 and durum=="Çıkış") else (KOYU_YESIL if (j==11) else SIYAH)
            td(ws2,row,j+1,v,bg=bg,renk=renk,sag=(j==7),bold=(j==11))

    # ════ SEKME 3: LOT BAZLI ════
    ws3 = wb.create_sheet("Lot Bazlı")
    ws3.sheet_view.showGridLines = False
    s3=[("Lot/Parti No",20),("SID",14),("Ürün Adı",32),("SKT",12),
        ("Adet",12),("Toplam Ağırlık (kg)",20),("Depo",26),("Depo Kodu",14)]
    for j,(b,w) in enumerate(s3):
        ws3.column_dimensions[get_column_letter(j+1)].width=w
        th(ws3,1,j+1,b,bg=KOYU_YESIL)
    ws3.row_dimensions[1].height=18; ws3.freeze_panes="A2"

    for ri,l in enumerate(veri["lot_dagilim"]):
        row=ri+2; ws3.row_dimensions[row].height=14
        bg=ACIK_GRI if ri%2==1 else BEYAZ
        vals=[l["lot"],l["sid"],l["urun_adi"],l["skt"],
              l["adet"],l["agirlik"],l["depo_adi"],l["depo_kod"]]
        for j,v in enumerate(vals):
            td(ws3,row,j+1,v,bg=bg,sag=(j in [4,5]))

    # ════ SEKME 4: HAREKETLER ════
    ws4 = wb.create_sheet("Hareketler")
    ws4.sheet_view.showGridLines = False
    s4=[("Tarih",20),("Tür",10),("Depo",26),("Belge No",20),
        ("Gönderen/Sebep",24),("Ürün Sayısı",14),("Çeşit/—",10),("Ağırlık (kg)",16)]
    for j,(b,w) in enumerate(s4):
        ws4.column_dimensions[get_column_letter(j+1)].width=w
        th(ws4,1,j+1,b,bg=TURUNCU)
    ws4.row_dimensions[1].height=18; ws4.freeze_panes="A2"

    ri=0
    for g in veri["son_girisler"]:
        row=ri+2; ws4.row_dimensions[row].height=14
        vals=[(g["tarih"]or"").replace("T"," "),"GİRİŞ",
              g.get("depo_adi",""),g.get("gonderi_no",""),g.get("gonderen",""),
              g.get("urun_sayisi",0),g.get("cesit_sayisi",0),g.get("toplam_agirlik",0)]
        for j,v in enumerate(vals):
            td(ws4,row,j+1,v,bg=ACIK_YESIL,renk=KOYU_YESIL if j==1 else SIYAH,
               sag=(j in[5,6,7]),bold=(j==1))
        ri+=1

    for c in veri["son_cikislar"]:
        row=ri+2; ws4.row_dimensions[row].height=14
        vals=[(c["tarih"]or"").replace("T"," "),"ÇIKIŞ",
              c.get("depo_adi",""),c.get("cikis_no",""),c.get("sebep",""),
              c.get("urun_sayisi",0),"—",c.get("toplam_agirlik",0)]
        for j,v in enumerate(vals):
            td(ws4,row,j+1,v,bg=ACIK_KIRM,renk=KIRMIZI if j==1 else SIYAH,
               sag=(j in[5,7]),bold=(j==1))
        ri+=1

    if ri==0:
        ws4.merge_cells("A2:H2")
        c=ws4.cell(row=2,column=1,value="Henüz hareket kaydı bulunmamaktadır.")
        c.font=Font(italic=True,color=GRI,size=9)
        c.alignment=Alignment(horizontal="center")

    # ── Kaydet ──
    zaman = datetime.now().strftime("%Y%m%d_%H%M%S")
    dosya_adi = f"PATBIS_Rapor_{zaman}.xlsx"
    masaustu = os.path.join(os.path.expanduser("~"), "Desktop")
    if not os.path.isdir(masaustu):
        masaustu = os.path.expanduser("~")
    hedef = os.path.join(masaustu, dosya_adi)
    try:
        wb.save(hedef)
    except Exception as e1:
        try:
            db_dir = os.path.dirname(_db_yolu())
            hedef = os.path.join(db_dir, dosya_adi)
            wb.save(hedef)
        except Exception as e2:
            return {"basarili": False, "mesaj": f"Dosya kaydedilemedi. Masaüstü hatası: {e1} | Yedek dizin hatası: {e2}"}

    return {
        "basarili": True,
        "dosya": hedef,
        "dosya_adi": dosya_adi,
        "mesaj": f"✅ Excel raporu oluşturuldu: {dosya_adi}",
        "istatistik": {"urun": len(urun_uniq), "lot": len(veri["lot_dagilim"])},
    }


# ══════════════════════════════════════════════════════════════════════════════
# DOSYA İŞLEMLERİ
# ══════════════════════════════════════════════════════════════════════════════

def _dosya_kaydet(params):
    """Metin içeriğini Masaüstüne kaydeder."""
    dosya_adi = params.get("dosya_adi")
    icerik = params.get("icerik")
    if not dosya_adi or not icerik:
        return {"basarili": False, "mesaj": "Dosya adı veya içerik eksik."}
        
    try:
        masaustu = os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.isdir(masaustu):
            masaustu = os.path.join(os.path.expanduser("~"), "OneDrive", "Masaüstü")
        if not os.path.isdir(masaustu):
            masaustu = os.path.expanduser("~")
            
        hedef = os.path.join(masaustu, dosya_adi)
        with open(hedef, "w", encoding="utf-8") as f:
            f.write(icerik)
            
        return {"basarili": True, "mesaj": f"Kaydedildi: {hedef}"}
    except Exception as e:
        return {"basarili": False, "mesaj": f"Dosya kaydedilemedi: {e}"}


# ══════════════════════════════════════════════════════════════════════════════
# YARDIMCI
# ══════════════════════════════════════════════════════════════════════════════

def _text(element, tag):
    """XML element'ten text çek, None-safe."""
    el = element.find(tag)
    return el.text.strip() if el is not None and el.text else ""


def _tarih_temizle(tarih):
    """Timezone offset'i temizle: 2025-04-16+03:00 → 2025-04-16"""
    if tarih and "+" in tarih:
        return tarih.split("+")[0]
    if tarih and "T" in tarih:
        return tarih.split("T")[0]
    return tarih


# ══════════════════════════════════════════════════════════════════════════════
# ANA DÖNGÜ
# ══════════════════════════════════════════════════════════════════════════════

def _islem_gunlugu_getir(params):
    try:
        conn = _db_baglanti()
        rows = conn.execute("SELECT * FROM islem_gunlugu ORDER BY id DESC LIMIT ?", (params.get("limit", 100),)).fetchall()
        conn.close()
        return {"basarili": True, "loglar": [dict(r) for r in rows]}
    except Exception as e:
        return {"basarili": False, "mesaj": str(e)}

def _urun_sorunlari_getir(params):
    try:
        conn = _db_baglanti()
        rows = conn.execute("SELECT * FROM urun_sorunlari ORDER BY id DESC").fetchall()
        conn.close()
        return {"basarili": True, "sorunlar": [dict(r) for r in rows]}
    except Exception as e:
        return {"basarili": False, "mesaj": str(e)}

def _sunucu_durumu_getir():
    try:
        from core import api_server
        ips = api_server.get_local_ips()
        port = api_server.PORT
        # Join all IPs into the dual-IP sync format
        qr_endpoints = [f"http://{ip}:{port}" for ip in ips]
        if not qr_endpoints:
            qr_endpoints = [f"http://127.0.0.1:{port}"]
        qr_data = "PATBIS_SYNC|" + "|".join(qr_endpoints)
        return {
            "basarili": True,
            "ips": ips,
            "port": port,
            "qr_data": qr_data
        }
    except Exception as e:
        return {"basarili": False, "mesaj": str(e)}

def main():
    # Windows'ta stdout/stdin varsayılan encoding cp1254 olabiliyor
    # Rust tarafı UTF-8 bekliyor — zorla
    if os.name == "nt":
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
        sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding="utf-8")

    # Sunucu API'sini arka planda (Daemon Thread olarak) başlat!
    import threading
    try:
        from core import api_server
        t = threading.Thread(target=api_server.main, daemon=True)
        t.start()
    except Exception as e:
        pass

    startup()

    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        try:
            req = json.loads(line)
            resp = handle(req)
            print(json.dumps(resp, ensure_ascii=False), flush=True)
        except json.JSONDecodeError:
            err = {"id": 0, "result": None, "error": "Geçersiz JSON"}
            print(json.dumps(err), flush=True)


# ── Google Drive İşlemleri ──

def _drive_durum():
    """Drive bağlantı durumunu ve son yedek bilgisini döndür."""
    try:
        config = drive_manager.config_yukle()
        bagli = drive_manager.drive_bagli_mi()
        return {
            "basarili": True,
            "bagli": bagli,
            "klasor": config.get("drive_folder_name"),
            "son_yedek": config.get("son_yedek_tarihi"),
            "son_dosya": config.get("son_yedek_dosya"),
            "hata": config.get("son_drive_hata"),
            "drive_available": drive_manager.DRIVE_AVAILABLE
        }
    except Exception as e:
        return {"basarili": False, "mesaj": str(e)}

def _drive_baglan():
    """OAuth2 akışını başlat."""
    try:
        service = drive_manager.drive_baglanti()
        if service:
            return {"basarili": True, "mesaj": "Google Drive bağlantısı başarıyla kuruldu."}
        return {"basarili": False, "mesaj": "Bağlantı kurulamadı."}
    except Exception as e:
        return {"basarili": False, "mesaj": f"Bağlantı hatası: {e}"}

def _drive_yedekle():
    """Veritabanını Drive'a yükle."""
    try:
        db_path = _db_yolu()
        sonuc = drive_manager.dosya_yukle(db_path, f"Manuel yedek — {datetime.now():%d.%m.%Y %H:%M}")
        
        # Config güncelle
        config = drive_manager.config_yukle()
        config["son_yedek_tarihi"] = str(datetime.now().date())
        config["son_yedek_dosya"] = sonuc.get("name")
        drive_manager.config_kaydet(config)
        
        return {"basarili": True, "mesaj": "Yedekleme başarıyla tamamlandı.", "dosya": sonuc.get("name")}
    except Exception as e:
        return {"basarili": False, "mesaj": f"Yedekleme hatası: {e}"}

def _drive_dosyalar():
    """Drive'daki yedekleri listele."""
    try:
        dosyalar = drive_manager.drive_dosyalari_listele(limit=15)
        return {"basarili": True, "dosyalar": dosyalar}
    except Exception as e:
        return {"basarili": False, "mesaj": f"Liste hatası: {e}"}

def _drive_login(params):
    """Şifre kontrolü yap."""
    try:
        raw_sifre = params.get("sifre", "")
        personel_id = params.get("personel_id", None)
        # Boşlukları temizle ve büyük harfe çevir (İ -> I dönüşümü dahil)
        sifre = raw_sifre.strip().upper().replace("İ", "I")
        clean_sifre = "".join(c for c in sifre if c.isalnum())
        
        # Admin bypass
        if clean_sifre == "GAMALI2026" or clean_sifre == "GAMALİ2026":
            return {"basarili": True, "mesaj": "Admin Girişi başarılı.", "kullanici": {"id": 0, "ad_soyad": "Admin", "unvan": "Yönetici"}}
        
        # Personel girişi
        if personel_id:
            conn = _db_baglanti()
            user = conn.execute("SELECT * FROM personeller WHERE id=? AND aktif=1", (personel_id,)).fetchone()
            conn.close()
            
            if user:
                db_sifre = user["sifre"] or ""
                # Basit şifre karşılaştırma
                if db_sifre.strip() == raw_sifre.strip():
                    return {"basarili": True, "mesaj": f"Giriş başarılı: {user['ad_soyad']}", "kullanici": dict(user)}
                else:
                    return {"basarili": False, "mesaj": "Hatalı şifre!"}
            else:
                return {"basarili": False, "mesaj": "Kullanıcı bulunamadı!"}

        # Eğer personel_id yoksa, drive_manager'dan yedek kontrolü yap
        if drive_manager.sifre_kontrol(sifre):
            return {"basarili": True, "mesaj": "Giriş başarılı.", "kullanici": {"id": 0, "ad_soyad": "Sistem", "unvan": ""}}
            
        return {"basarili": False, "mesaj": f"Hatalı şifre!"}
    except Exception as e:
        return {"basarili": False, "mesaj": f"Giriş hatası: {str(e)}"}


# ── Personel İşlemleri ──

def _personeller_getir():
    """Tüm personelleri listele."""
    try:
        conn = _db_baglanti()
        rows = conn.execute("SELECT * FROM personeller WHERE aktif=1 ORDER BY ad_soyad").fetchall()
        result = [dict(r) for r in rows]
        conn.close()
        return {"basarili": True, "personeller": result}
    except Exception as e:
        return {"basarili": False, "mesaj": str(e)}

def _personel_ekle(params):
    """Yeni personel ekle."""
    try:
        ad_soyad = params.get("ad_soyad")
        unvan = params.get("unvan", "")
        sifre = params.get("sifre", "")
        if not ad_soyad:
            return {"basarili": False, "mesaj": "Ad soyad gerekli"}
        
        conn = _db_baglanti()
        conn.execute("INSERT INTO personeller (ad_soyad, unvan, sifre) VALUES (?, ?, ?)", (ad_soyad, unvan, sifre))
        conn.commit()
        conn.close()
        return {"basarili": True, "mesaj": f"Personel eklendi: {ad_soyad}"}
    except Exception as e:
        return {"basarili": False, "mesaj": str(e)}

def _personel_guncelle(params):
    """Personel bilgilerini güncelle."""
    try:
        pid = params.get("id")
        ad_soyad = params.get("ad_soyad")
        unvan = params.get("unvan")
        sifre = params.get("sifre", "")
        if not pid:
            return {"basarili": False, "mesaj": "ID gerekli"}
        
        conn = _db_baglanti()
        if sifre:
            conn.execute("UPDATE personeller SET ad_soyad=?, unvan=?, sifre=? WHERE id=?", (ad_soyad, unvan, sifre, pid))
        else:
            conn.execute("UPDATE personeller SET ad_soyad=?, unvan=? WHERE id=?", (ad_soyad, unvan, pid))
        conn.commit()
        conn.close()
        return {"basarili": True, "mesaj": "Personel güncellendi."}
    except Exception as e:
        return {"basarili": False, "mesaj": str(e)}

def _personel_sil(params):
    """Personeli (yumuşak) sil."""
    try:
        pid = params.get("id")
        if not pid:
            return {"basarili": False, "mesaj": "ID gerekli"}
        
        conn = _db_baglanti()
        conn.execute("UPDATE personeller SET aktif=0 WHERE id=?", (pid,))
        conn.commit()
        conn.close()
        return {"basarili": True, "mesaj": "Personel silindi."}
    except Exception as e:
        return {"basarili": False, "mesaj": str(e)}


if __name__ == "__main__":
    main()