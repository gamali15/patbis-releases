"""
GAMALI NEXUS PATBiS — Excel Rapor Uretici
openpyxl ile 4 sekmeli stok raporu olusturur.
"""
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# ── RENK PALETİ (GAMALI Cyberpunk) ──────────────────────────────────────────
C_HEADER_BG   = "1A1A2E"   # Koyu lacivert — başlık satırı
C_HEADER_FG   = "4CAF50"   # GAMALI Yeşil — başlık yazısı
C_SUBHDR_BG   = "16213E"   # Bölüm başlığı arka plan
C_SUBHDR_FG   = "E8A838"   # ALP Gold — bölüm başlıkları
C_ROW_ALT     = "0F3460"   # Satır alt rengi
C_WHITE       = "FFFFFF"
C_RED_BG      = "C0392B"   # SKT uyarısı arka plan
C_RED_FG      = "FFFFFF"
C_ORANGE_BG   = "E67E22"   # SKT yaklaşıyor
C_GREEN_BG    = "1E8449"   # Stokta — iyi
C_DARK_BG     = "0D1117"   # Satır arka plan (koyu)

# ── STİL YARDIMCILARI ────────────────────────────────────────────────────────
def _font(bold=False, color=C_WHITE, size=10, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size)

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _border():
    s = Side(style="thin", color="2C2C54")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def _header_row(ws, row, cols, bg=C_HEADER_BG, fg=C_HEADER_FG):
    """Başlık satırı stilini uygula"""
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=True, color=fg, size=10)
        c.fill      = _fill(bg)
        c.alignment = _center()
        c.border    = _border()

def _data_row(ws, row, values, alt=False, bold=False):
    """Veri satırı stilini uygula"""
    bg = C_ROW_ALT if alt else C_DARK_BG
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=bold, color=C_WHITE, size=9)
        c.fill      = _fill(bg)
        c.alignment = _left()
        c.border    = _border()

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _title_block(ws, baslik, alt_baslik, tarih_str):
    """Sayfanın en üstüne başlık bloğu ekle"""
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 14
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = baslik
    c.font      = _font(bold=True, color=C_HEADER_FG, size=14, name="Arial")
    c.fill      = _fill(C_HEADER_BG)
    c.alignment = _center()
    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value     = alt_baslik
    c2.font      = _font(bold=False, color=C_SUBHDR_FG, size=10)
    c2.fill      = _fill(C_SUBHDR_BG)
    c2.alignment = _center()
    ws.merge_cells("A3:H3")
    c3 = ws["A3"]
    c3.value     = f"Rapor Tarihi: {tarih_str}"
    c3.font      = _font(color="888888", size=9)
    c3.fill      = _fill(C_DARK_BG)
    c3.alignment = _center()


# ── SEKME 1: ÖZET ────────────────────────────────────────────────────────────
def _sekme_ozet(wb, depolar_ozet, tarih_str):
    ws = wb.active
    ws.title = "📊 ÖZET"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    _title_block(ws, "GAMALI NEXUS PATBİS — STOK ÖZET RAPORU",
                 "Tüm Depolar Genel Bakış", tarih_str)

    hdrs = ["DEPO KODU", "DEPO ADI", "STOKTA ÜRÜN", "ÇIKTI ÜRÜN",
            "TOPLAM ÜRÜN", "PALET", "KOLİ", "DOLULUK %"]
    _header_row(ws, 4, hdrs)

    toplam_stokta = toplam_cikti = toplam_palet = toplam_koli = 0
    for i, d in enumerate(depolar_ozet):
        stokta  = d.get("stokta") or 0
        cikti   = d.get("cikti")  or 0
        toplam  = stokta + cikti
        palet   = d.get("palet_sayisi") or 0
        koli    = d.get("kutu_sayisi")  or 0
        doluluk = f"=IFERROR(C{i+5}/E{i+5},0)"
        row_vals = [
            d.get("kod",""), d.get("ad",""),
            stokta, cikti, toplam, palet, koli, doluluk
        ]
        _data_row(ws, i+5, row_vals, alt=(i%2==1))
        # Doluluk hücresi yüzde formatı
        ws.cell(i+5, 8).number_format = "0.0%"
        toplam_stokta += stokta
        toplam_cikti  += cikti
        toplam_palet  += palet
        toplam_koli   += koli

    # Toplam satırı
    tr = len(depolar_ozet) + 5
    ws.row_dimensions[tr].height = 20
    toplam_vals = ["", "GENEL TOPLAM", toplam_stokta, toplam_cikti,
                   toplam_stokta+toplam_cikti, toplam_palet, toplam_koli, ""]
    _header_row(ws, tr, toplam_vals, bg=C_SUBHDR_BG, fg=C_SUBHDR_FG)

    _col_widths(ws, [16, 28, 14, 12, 14, 10, 10, 12])

    for r in range(1, tr+1):
        ws.row_dimensions[r].height = ws.row_dimensions[r].height or 16


# ── SEKME 2: STOK DETAY ──────────────────────────────────────────────────────
def _sekme_stok_detay(wb, urunler, tarih_str):
    ws = wb.create_sheet("📦 STOK DETAY")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    _title_block(ws, "GAMALI NEXUS PATBİS — STOK DETAY",
                 "Stokta Bulunan Tüm Ürünler", tarih_str)

    # Filtre satırı (depo göstergesi)
    ws.merge_cells("A4:L4")
    c4 = ws["A4"]
    c4.value     = f"Toplam {len(urunler)} ürün listelenmektedir."
    c4.font      = _font(color=C_SUBHDR_FG, size=9)
    c4.fill      = _fill(C_SUBHDR_BG)
    c4.alignment = _left()

    hdrs = ["DEPO", "UID", "PSN", "ÜRÜN ADI", "ÜRÜN KODU",
            "BATCH", "SKT", "ÜRETİM TARİHİ", "AĞIRLIK NET",
            "AĞIRLIK TİP", "PALET UID", "KOLİ UID"]
    _header_row(ws, 5, hdrs)

    bugun = datetime.today().date()
    for i, u in enumerate(urunler):
        skt_str = u.get("skt","")
        row_vals = [
            u.get("depo_kod",""),
            u.get("uid",""),
            u.get("psn",""),
            u.get("urun_adi",""),
            u.get("urun_kodu",""),
            u.get("batch",""),
            skt_str,
            u.get("uretim_tarihi",""),
            u.get("agirlik_net",""),
            u.get("agirlik_tip",""),
            u.get("palet_uid",""),
            u.get("kutu_uid",""),
        ]
        alt = (i % 2 == 1)
        _data_row(ws, i+6, row_vals, alt=alt)

        # SKT renklendir
        if skt_str:
            try:
                skt_date = datetime.strptime(skt_str[:10], "%Y-%m-%d").date()
                kalan = (skt_date - bugun).days
                skt_cell = ws.cell(i+6, 7)
                if kalan <= 30:
                    skt_cell.fill = _fill(C_RED_BG)
                    skt_cell.font = _font(bold=True, color=C_WHITE)
                elif kalan <= 90:
                    skt_cell.fill = _fill(C_ORANGE_BG)
                    skt_cell.font = _font(bold=True, color=C_WHITE)
            except:
                pass

    _col_widths(ws, [16, 14, 8, 24, 12, 14, 12, 14, 12, 12, 14, 14])
    ws.auto_filter.ref = f"A5:L5"


# ── SEKME 3: ÇIKIŞ GEÇMİŞİ ──────────────────────────────────────────────────
def _sekme_cikis(wb, cikislar, tarih_str):
    ws = wb.create_sheet("🚚 ÇIKIŞ GEÇMİŞİ")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    _title_block(ws, "GAMALI NEXUS PATBİS — ÇIKIŞ GEÇMİŞİ",
                 "Son 200 Stok Çıkış Kaydı", tarih_str)

    ws.merge_cells("A4:J4")
    c4 = ws["A4"]
    c4.value     = f"Toplam {len(cikislar)} çıkış kaydı listelenmektedir."
    c4.font      = _font(color=C_SUBHDR_FG, size=9)
    c4.fill      = _fill(C_SUBHDR_BG)
    c4.alignment = _left()

    hdrs = ["DEPO", "UID", "PSN", "ÜRÜN ADI", "ÜRÜN KODU",
            "BATCH", "SKT", "AĞIRLIK NET", "ÇIKIŞ TARİHİ", "AĞIRLIK TİP"]
    _header_row(ws, 5, hdrs)

    for i, c in enumerate(cikislar):
        row_vals = [
            c.get("depo_kod",""),
            c.get("uid",""),
            c.get("psn",""),
            c.get("urun_adi",""),
            c.get("urun_kodu",""),
            c.get("batch",""),
            c.get("skt",""),
            c.get("agirlik_net",""),
            c.get("cikis_tarihi","")[:16] if c.get("cikis_tarihi") else "",
            c.get("agirlik_tip",""),
        ]
        _data_row(ws, i+6, row_vals, alt=(i%2==1))

    _col_widths(ws, [16, 14, 8, 24, 12, 14, 12, 12, 18, 12])
    ws.auto_filter.ref = "A5:J5"


# ── SEKME 4: SKT UYARISI ─────────────────────────────────────────────────────
def _sekme_skt(wb, skt_urunler, tarih_str):
    ws = wb.create_sheet("⚠️ SKT UYARI")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    _title_block(ws, "GAMALI NEXUS PATBİS — SKT UYARI LİSTESİ",
                 "Son Kullanma Tarihi 90 Gün İçinde Dolan Ürünler", tarih_str)

    ws.merge_cells("A4:I4")
    c4 = ws["A4"]
    kritik = sum(1 for u in skt_urunler
                 if u.get("skt") and
                    (datetime.strptime(u["skt"][:10],"%Y-%m-%d").date()
                     - datetime.today().date()).days <= 30)
    c4.value = (f"⚠️  {len(skt_urunler)} ürünün SKT'si 90 gün içinde dolmaktadır. "
                f"🔴 {kritik} ürün KRİTİK (≤30 gün).")
    c4.font      = _font(bold=True, color="FFFFFF", size=10)
    c4.fill      = _fill(C_RED_BG if kritik else "E67E22")
    c4.alignment = _left()

    hdrs = ["DEPO", "UID", "PSN", "ÜRÜN ADI", "BATCH",
            "SKT", "KALAN GÜN", "AĞIRLIK NET", "DURUM"]
    _header_row(ws, 5, hdrs)

    bugun = datetime.today().date()
    for i, u in enumerate(skt_urunler):
        skt_str = u.get("skt","")
        kalan = ""
        durum = ""
        try:
            skt_date = datetime.strptime(skt_str[:10], "%Y-%m-%d").date()
            kalan = (skt_date - bugun).days
            if kalan <= 30:
                durum = "🔴 KRİTİK"
            elif kalan <= 60:
                durum = "🟠 UYARI"
            else:
                durum = "🟡 TAKİPTE"
        except:
            pass

        row_vals = [
            u.get("depo_kod",""), u.get("uid",""), u.get("psn",""),
            u.get("urun_adi",""), u.get("batch",""),
            skt_str, kalan, u.get("agirlik_net",""), durum
        ]
        _data_row(ws, i+6, row_vals, alt=(i%2==1))

        # Satır renklendir
        if isinstance(kalan, int):
            if kalan <= 30:
                bg = C_RED_BG
            elif kalan <= 60:
                bg = "784212"
            else:
                bg = "7D6608"
            for col in range(1, 10):
                cell = ws.cell(i+6, col)
                cell.fill = _fill(bg)
                cell.font = _font(color=C_WHITE, size=9,
                                  bold=(kalan <= 30))

    _col_widths(ws, [16, 14, 8, 24, 14, 12, 12, 12, 12])
    ws.auto_filter.ref = "A5:I5"


# ── ANA FONKSİYON ─────────────────────────────────────────────────────────────
def rapor_olustur(cikis_klasoru: str, db_manager) -> str:
    """
    Excel raporu oluşturur ve dosya yolunu döner.
    db_manager: core.db_manager modülü (inject edilir)
    """
    tarih_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    dosya_adi = f"PATBIS_Rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    cikis_yolu = str(Path(cikis_klasoru) / dosya_adi)

    # Veri çek
    depolar_ozet = db_manager.depo_ozeti()

    # Stokta olan tüm ürünleri çek (detay için özel sorgu)
    import sqlite3
    conn = db_manager.get_connection()
    urunler = [dict(r) for r in conn.execute("""
        SELECT u.depo_kod, u.uid, u.psn, u.urun_adi, u.urun_kodu,
               u.batch, u.skt, u.uretim_tarihi, u.agirlik_net, u.agirlik_tip,
               u.kutu_uid,
               k.palet_uid
        FROM urunler u
        LEFT JOIN kutular k ON k.uid=u.kutu_uid AND k.depo_kod=u.depo_kod
        WHERE u.durum='STOKTA'
        ORDER BY u.depo_kod, u.skt, u.uid
    """).fetchall()]
    cikislar  = db_manager.son_cikislar(200)
    skt_liste = db_manager.skt_yaklasan_urunler(90)
    conn.close()

    wb = Workbook()
    _sekme_ozet(wb, depolar_ozet, tarih_str)
    _sekme_stok_detay(wb, urunler, tarih_str)
    _sekme_cikis(wb, cikislar, tarih_str)
    _sekme_skt(wb, skt_liste, tarih_str)

    wb.save(cikis_yolu)
    return cikis_yolu
